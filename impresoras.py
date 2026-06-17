import requests
from bs4 import BeautifulSoup
import re
import json
import os
import sys
import logging
import sqlite3
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from contextlib import contextmanager
import numpy as np
from types import SimpleNamespace
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from tkinter.ttk import Progressbar, Treeview, Scrollbar, Spinbox, Combobox, Style
import tkinter.ttk as ttk
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import matplotlib.ticker as mticker
import matplotlib.dates as mdates

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

BASE_DIR         = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) \
                   else os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH      = os.path.join(BASE_DIR, "config.json")
LOG_PATH         = os.path.join(BASE_DIR, "errores.log")
_DB_PATH_DEFAULT = os.path.join(BASE_DIR, "impresoras.db")
DB_PATH          = _DB_PATH_DEFAULT   # puede actualizarse desde config al iniciar

COLOR_BAJO      = "#FF6B6B"
COLOR_MEDIO     = "#FFD93D"
COLOR_SIN_DATOS = "#CCCCCC"

BG_MAIN           = "#F5F5F5"
COLOR_ACCENT      = "#2C7BB6"
COLOR_ACCENT_DARK = "#1A5F8A"
BG_TREE_PAR       = "#FFFFFF"
BG_TREE_IMPAR     = "#EBF4FB"
FONT_UI           = ("Segoe UI", 9)
FONT_BOLD         = ("Segoe UI", 9, "bold")
FONT_TABLE        = ("Segoe UI", 9)
FONT_NUM          = ("Segoe UI", 11, "bold")

# 6 columnas: Sucursal primero
COLUMNAS_TREE = ("Sucursal", "IP", "Modelo", "Último Monitoreo", "Tóner (%)", "Unidad Imagen (%)", "Kit Mantenimiento (%)")

TIPOS_INSUMO = ["Tóner", "Unidad Imagen"]

INTERVALOS_AUTO = {
    "15 min":   900,
    "30 min":   1800,
    "1 hora":   3600,
    "2 horas":  7200,
    "4 horas":  14400,
    "24 horas": 86400,
    "48 horas": 172800,
}

# Configuración de modelos: índices en la lista de porcentajes del HTML.
# Orden de cada lista: [tóner, kit_mantenimiento, unidad_imagen].
# None indica que ese consumible no existe en el modelo.
MODELOS_CONFIG = {
    "Lexmark MX611dhe": [0, 2, 3],
    "Lexmark X466de":   [0, 2, 3],
    "Lexmark X464de":   [0, 2, 3],
    "Lexmark MX710":    [0, 2, 4],
    "Lexmark MS811":    [0, 2, 4],
    "Lexmark MS812":    [0, 2, 4],
    "Lexmark T654":     [0, None, 2],
}

# ---------------------------------------------------------------------------
# Helpers de UI — estilo y tooltips
# ---------------------------------------------------------------------------

class Tooltip:
    """Muestra un pequeño popup de ayuda al pasar el cursor sobre un widget."""
    def __init__(self, widget, text):
        self._top = None
        widget.bind("<Enter>", lambda _: self._show(widget, text))
        widget.bind("<Leave>", lambda _: self._hide())

    def _show(self, widget, text):
        x = widget.winfo_rootx() + widget.winfo_width() // 2
        y = widget.winfo_rooty() + widget.winfo_height() + 4
        self._top = tk.Toplevel(widget)
        self._top.wm_overrideredirect(True)
        self._top.wm_geometry(f"+{x}+{y}")
        tk.Label(self._top, text=text, bg="#FFFFCC", fg="#333333",
                 relief="solid", borderwidth=1,
                 font=("Segoe UI", 8), padx=6, pady=3).pack()

    def _hide(self):
        if self._top:
            self._top.destroy()
            self._top = None


def _estilo_btn(btn, primario=True):
    """Aplica estilo plano con color de acento y efecto hover a un tk.Button."""
    bg    = COLOR_ACCENT      if primario else "#E0E0E0"
    fg    = "white"           if primario else "#333333"
    hover = COLOR_ACCENT_DARK if primario else "#C8C8C8"
    btn.config(bg=bg, fg=fg, font=FONT_UI, relief="flat",
               activebackground=hover, activeforeground=fg,
               cursor="hand2", padx=10, pady=5, bd=0)
    btn.bind("<Enter>", lambda _: btn.config(bg=hover))
    btn.bind("<Leave>", lambda _: btn.config(bg=bg))


# ---------------------------------------------------------------------------
# Log de errores
# ---------------------------------------------------------------------------

_log = logging.getLogger("impresoras")
_log.setLevel(logging.ERROR)
if not _log.handlers:
    _h = logging.FileHandler(LOG_PATH, encoding="utf-8")
    _h.setFormatter(logging.Formatter("%(asctime)s  %(message)s", datefmt="%d/%m/%Y %H:%M:%S"))
    _log.addHandler(_h)

# ---------------------------------------------------------------------------
# Config persistente
# ---------------------------------------------------------------------------

def cargar_config():
    try:
        with open(CONFIG_PATH, "r") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

def guardar_config(**kwargs):
    config = cargar_config()
    config.update(kwargs)
    with open(CONFIG_PATH, "w") as f:
        json.dump(config, f, indent=2)


def _inicializar_db_path():
    """Lee db_path de config.json y actualiza la variable global DB_PATH."""
    global DB_PATH
    ruta = cargar_config().get("db_path", "").strip()
    if ruta:
        carpeta = os.path.dirname(ruta)
        if os.path.isdir(carpeta):
            DB_PATH = ruta
            return
    DB_PATH = _DB_PATH_DEFAULT

# ---------------------------------------------------------------------------
# Base de datos SQLite
# ---------------------------------------------------------------------------

@contextmanager
def db_connect():
    """Context manager para conexiones SQLite con commit/rollback automático."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def init_db():
    """Crea las tablas si no existen. Se llama al arrancar la aplicación."""
    with db_connect() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS impresoras (
                id       INTEGER PRIMARY KEY AUTOINCREMENT,
                ip       TEXT NOT NULL UNIQUE,
                modelo   TEXT NOT NULL,
                sucursal TEXT NOT NULL DEFAULT '',
                nombre   TEXT          DEFAULT '',
                sn       TEXT          DEFAULT '',
                activa   INTEGER NOT NULL DEFAULT 1
            );
            CREATE TABLE IF NOT EXISTS monitoreos (
                id                INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha             TEXT NOT NULL,
                ip                TEXT NOT NULL,
                toner             REAL,
                unidad_imagen     REAL,
                kit_mantenimiento REAL
            );
            CREATE TABLE IF NOT EXISTS envios (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha            TEXT NOT NULL,
                sucursal         TEXT NOT NULL,
                ip               TEXT DEFAULT '',
                tipo_insumo      TEXT NOT NULL,
                modelo_impresora TEXT NOT NULL,
                cantidad         INTEGER NOT NULL DEFAULT 1
            );
        """)
        # Migración: agrega columna sn a DBs existentes que no la tienen
        cols = {r[1] for r in conn.execute("PRAGMA table_info(impresoras)").fetchall()}
        if "sn" not in cols:
            conn.execute("ALTER TABLE impresoras ADD COLUMN sn TEXT DEFAULT ''")
        # Migración: tablas de stock (para DBs creadas antes de esta versión)
        tablas = {r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
        if "stock_deposito" not in tablas:
            conn.execute("""
                CREATE TABLE stock_deposito (
                    id               INTEGER PRIMARY KEY AUTOINCREMENT,
                    tipo_insumo      TEXT    NOT NULL,
                    modelo_impresora TEXT    NOT NULL,
                    cantidad         INTEGER NOT NULL DEFAULT 0,
                    stock_minimo     INTEGER NOT NULL DEFAULT 2,
                    UNIQUE(tipo_insumo, modelo_impresora)
                )
            """)
        if "movimientos_stock" not in tablas:
            conn.execute("""
                CREATE TABLE movimientos_stock (
                    id               INTEGER PRIMARY KEY AUTOINCREMENT,
                    fecha            TEXT    NOT NULL,
                    tipo             TEXT    NOT NULL,
                    tipo_insumo      TEXT    NOT NULL,
                    modelo_impresora TEXT    NOT NULL,
                    cantidad         INTEGER NOT NULL,
                    observacion      TEXT    DEFAULT '',
                    envio_id         INTEGER DEFAULT NULL
                )
            """)
        # Migración: columna anulado en envios
        cols_env = {r[1] for r in conn.execute("PRAGMA table_info(envios)").fetchall()}
        if "anulado" not in cols_env:
            conn.execute("ALTER TABLE envios ADD COLUMN anulado INTEGER NOT NULL DEFAULT 0")
        # Migración: columna ubicacion en impresoras
        if "ubicacion" not in cols:
            conn.execute("ALTER TABLE impresoras ADD COLUMN ubicacion TEXT DEFAULT ''")
        # Migración: tabla modelos
        tablas2 = {r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
        if "modelos" not in tablas2:
            conn.execute("""
                CREATE TABLE modelos (
                    id     INTEGER PRIMARY KEY AUTOINCREMENT,
                    nombre TEXT NOT NULL UNIQUE
                )
            """)
            # Poblar desde MODELOS_CONFIG + modelos existentes en impresoras
            conocidos = set()
            for m in MODELOS_CONFIG:
                conn.execute("INSERT OR IGNORE INTO modelos (nombre) VALUES (?)", (m,))
                conocidos.add(m)
            extras = conn.execute(
                "SELECT DISTINCT modelo FROM impresoras WHERE modelo!=''").fetchall()
            for (m,) in extras:
                if m not in conocidos:
                    conn.execute("INSERT OR IGNORE INTO modelos (nombre) VALUES (?)", (m,))
        # Migración: columna modelo_id en impresoras
        cols3 = {r[1] for r in conn.execute("PRAGMA table_info(impresoras)").fetchall()}
        if "modelo_id" not in cols3:
            conn.execute("ALTER TABLE impresoras ADD COLUMN modelo_id INTEGER REFERENCES modelos(id)")
            conn.execute("""
                UPDATE impresoras SET modelo_id = (
                    SELECT id FROM modelos WHERE modelos.nombre = impresoras.modelo
                )
            """)


def db_impresoras_todas(activas_solo=False):
    """Retorna lista de dicts con todas las impresoras."""
    with db_connect() as conn:
        q = "SELECT * FROM impresoras"
        if activas_solo:
            q += " WHERE activa = 1"
        q += " ORDER BY sucursal, ip"
        return [dict(r) for r in conn.execute(q).fetchall()]


def db_impresora_agregar(ip, modelo, sucursal, nombre="", sn="", ubicacion=""):
    modelo_id = _modelo_id_obtener_o_crear(modelo)
    with db_connect() as conn:
        conn.execute(
            "INSERT INTO impresoras (ip, modelo, sucursal, nombre, sn, ubicacion, modelo_id) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (ip, modelo, sucursal, nombre, sn, ubicacion, modelo_id),
        )


def db_impresora_actualizar(id_, ip, modelo, sucursal, nombre, sn, activa, ubicacion=""):
    modelo_id = _modelo_id_obtener_o_crear(modelo)
    with db_connect() as conn:
        conn.execute(
            "UPDATE impresoras SET ip=?, modelo=?, sucursal=?, nombre=?, sn=?, activa=?, ubicacion=?, modelo_id=? WHERE id=?",
            (ip, modelo, sucursal, nombre, sn, 1 if activa else 0, ubicacion, modelo_id, id_),
        )


def db_impresora_eliminar(id_):
    with db_connect() as conn:
        conn.execute("DELETE FROM impresoras WHERE id=?", (id_,))


def db_ultimo_toner(ip):
    """Retorna el último nivel de tóner registrado para una IP, o None."""
    with db_connect() as conn:
        row = conn.execute(
            "SELECT toner FROM monitoreos WHERE ip=? AND toner IS NOT NULL ORDER BY fecha DESC LIMIT 1",
            (ip,)).fetchone()
        return row[0] if row else None


def db_modelos_activos():
    """Retorna lista ordenada de modelos únicos de impresoras activas."""
    with db_connect() as conn:
        rows = conn.execute(
            "SELECT DISTINCT modelo FROM impresoras WHERE activa=1 AND modelo!='' ORDER BY modelo").fetchall()
        return [r[0] for r in rows]


def db_sucursales_activas():
    """Retorna lista ordenada de sucursales únicas de impresoras activas."""
    with db_connect() as conn:
        rows = conn.execute(
            "SELECT DISTINCT sucursal FROM impresoras WHERE activa=1 AND sucursal!='' ORDER BY sucursal").fetchall()
        return [r[0] for r in rows]


def _modelo_id_obtener_o_crear(nombre):
    """Obtiene el id del modelo, creándolo si no existe en la tabla modelos."""
    with db_connect() as conn:
        conn.execute("INSERT OR IGNORE INTO modelos (nombre) VALUES (?)", (nombre.strip(),))
        row = conn.execute("SELECT id FROM modelos WHERE nombre=?", (nombre.strip(),)).fetchone()
        return row["id"] if row else None


def db_modelos_listar():
    """Retorna lista de (id, nombre) de todos los modelos ordenados."""
    with db_connect() as conn:
        rows = conn.execute("SELECT id, nombre FROM modelos ORDER BY nombre").fetchall()
        return [(r["id"], r["nombre"]) for r in rows]


def db_modelo_agregar(nombre):
    """Inserta un modelo. Retorna el id. Lanza IntegrityError si ya existe."""
    with db_connect() as conn:
        cur = conn.execute("INSERT INTO modelos (nombre) VALUES (?)", (nombre.strip(),))
        return cur.lastrowid


def db_modelo_renombrar(id_, nuevo_nombre):
    """Renombra un modelo y actualiza el texto en todas las impresoras que lo usan."""
    nuevo = nuevo_nombre.strip()
    with db_connect() as conn:
        conn.execute("UPDATE modelos SET nombre=? WHERE id=?", (nuevo, id_))
        conn.execute("UPDATE impresoras SET modelo=? WHERE modelo_id=?", (nuevo, id_))


def db_modelo_eliminar(id_):
    """Elimina un modelo solo si ninguna impresora lo referencia."""
    with db_connect() as conn:
        usado = conn.execute(
            "SELECT COUNT(*) FROM impresoras WHERE modelo_id=?", (id_,)).fetchone()[0]
        if usado > 0:
            raise ValueError(f"No se puede eliminar: {usado} impresora(s) usan este modelo.")
        conn.execute("DELETE FROM modelos WHERE id=?", (id_,))


def db_registrar_envio(fecha, sucursal, ip, tipo, modelo, cantidad):
    """Inserta un envío de insumo y descuenta del stock, creando entrada si no existe."""
    with db_connect() as conn:
        cur = conn.execute(
            "INSERT INTO envios (fecha, sucursal, ip, tipo_insumo, modelo_impresora, cantidad) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (fecha.strftime("%Y-%m-%d %H:%M:%S"), sucursal, ip or "", tipo, modelo, cantidad),
        )
        envio_id = cur.lastrowid
        conn.execute(
            "INSERT OR IGNORE INTO stock_deposito (tipo_insumo, modelo_impresora, cantidad) "
            "VALUES (?, ?, 0)", (tipo, modelo))
        existe = conn.execute(
            "SELECT cantidad FROM stock_deposito WHERE tipo_insumo=? AND modelo_impresora=?",
            (tipo, modelo)).fetchone()
        nueva_cant = max(0, existe[0] - cantidad)
        conn.execute(
            "UPDATE stock_deposito SET cantidad=? WHERE tipo_insumo=? AND modelo_impresora=?",
            (nueva_cant, tipo, modelo))
        conn.execute(
            "INSERT INTO movimientos_stock "
            "(fecha, tipo, tipo_insumo, modelo_impresora, cantidad, observacion, envio_id) "
            "VALUES (?, 'salida', ?, ?, ?, ?, ?)",
            (fecha.strftime("%Y-%m-%d %H:%M:%S"), tipo, modelo, cantidad,
             f"Envío a {sucursal}", envio_id))


def db_stock_obtener():
    """Retorna lista de dicts con todo el stock del depósito."""
    with db_connect() as conn:
        return [dict(r) for r in conn.execute(
            "SELECT tipo_insumo, modelo_impresora, cantidad, stock_minimo "
            "FROM stock_deposito ORDER BY tipo_insumo, modelo_impresora").fetchall()]


def db_stock_agregar_entrada(tipo, modelo, cantidad, observacion=""):
    """Suma cantidad al stock del depósito y registra el movimiento de entrada."""
    with db_connect() as conn:
        conn.execute("""
            INSERT INTO stock_deposito (tipo_insumo, modelo_impresora, cantidad)
            VALUES (?, ?, ?)
            ON CONFLICT(tipo_insumo, modelo_impresora)
            DO UPDATE SET cantidad = cantidad + excluded.cantidad
        """, (tipo, modelo, cantidad))
        conn.execute("""
            INSERT INTO movimientos_stock
                (fecha, tipo, tipo_insumo, modelo_impresora, cantidad, observacion)
            VALUES (?, 'entrada', ?, ?, ?, ?)
        """, (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), tipo, modelo,
              cantidad, observacion))


def db_stock_editar_minimo(tipo, modelo, minimo):
    """Actualiza el stock mínimo de alerta para un tipo+modelo."""
    with db_connect() as conn:
        conn.execute(
            "UPDATE stock_deposito SET stock_minimo=? WHERE tipo_insumo=? AND modelo_impresora=?",
            (minimo, tipo, modelo))


def db_movimientos_stock(tipo_f="", modelo_f="", desde_dt=None, hasta_dt=None):
    """Retorna movimientos de stock filtrables."""
    with db_connect() as conn:
        q = ("SELECT fecha, tipo, tipo_insumo, modelo_impresora, cantidad, observacion "
             "FROM movimientos_stock WHERE 1=1")
        params = []
        if tipo_f:
            q += " AND tipo_insumo = ?";               params.append(tipo_f)
        if modelo_f:
            q += " AND LOWER(modelo_impresora) LIKE ?"; params.append(f"%{modelo_f.lower()}%")
        if desde_dt:
            q += " AND fecha >= ?";                    params.append(desde_dt.strftime("%Y-%m-%d"))
        if hasta_dt:
            q += " AND fecha <= ?";                    params.append(hasta_dt.strftime("%Y-%m-%d 23:59:59"))
        q += " ORDER BY fecha DESC"
        return conn.execute(q, params).fetchall()


def db_cargar_envios(sucursal_filtro="", anio=None, mes=0, filtro_anulado=None):
    """Retorna lista de tuplas (id, fecha, sucursal, ip, tipo, modelo, cantidad, anulado)."""
    with db_connect() as conn:
        q = ("SELECT id, fecha, sucursal, ip, tipo_insumo, modelo_impresora, cantidad, anulado "
             "FROM envios WHERE 1=1")
        params = []
        if sucursal_filtro:
            q += " AND LOWER(sucursal) LIKE ?"
            params.append(f"%{sucursal_filtro.lower()}%")
        if anio:
            q += " AND SUBSTR(fecha, 1, 4) = ?"
            params.append(str(anio))
        if mes:
            q += " AND SUBSTR(fecha, 6, 2) = ?"
            params.append(f"{mes:02d}")
        if filtro_anulado is True:
            q += " AND anulado = 1"
        elif filtro_anulado is False:
            q += " AND anulado = 0"
        q += " ORDER BY fecha DESC"
        rows = conn.execute(q, params).fetchall()

    filas = []
    for r in rows:
        try:
            fecha_dt  = datetime.strptime(r[1], "%Y-%m-%d %H:%M:%S")
            fecha_str = fecha_dt.strftime("%d/%m/%Y")
        except (ValueError, TypeError):
            fecha_str = str(r[1] or "")
        filas.append((r[0], fecha_str, r[2], r[3], r[4], r[5], r[6], r[7]))
    return filas

def db_cargar_historial(desde=None, hasta=None, sucursal="", modelo="", ip=""):
    """Retorna filas de monitoreos con JOIN a impresoras.
    desde/hasta: strings 'YYYY-MM-DD'. modelo/ip: filtro exacto (o vacío = sin filtro).
    Retorna lista de tuplas (7 datos + tag).
    """
    cfg          = cargar_config()
    umbral_bajo  = cfg.get("umbral_bajo",  10)
    umbral_medio = cfg.get("umbral_medio", 25)

    with db_connect() as conn:
        q = """
            SELECT m.fecha,
                   COALESCE(i.sucursal, '') AS sucursal,
                   m.ip,
                   COALESCE(i.modelo, '')   AS modelo,
                   m.toner, m.unidad_imagen, m.kit_mantenimiento
            FROM monitoreos m
            LEFT JOIN impresoras i ON m.ip = i.ip
            WHERE 1=1
        """
        params = []
        if desde:
            q += " AND m.fecha >= ?"
            params.append(desde + " 00:00:00")
        if hasta:
            q += " AND m.fecha <= ?"
            params.append(hasta + " 23:59:59")
        if sucursal and sucursal != "Todas":
            q += " AND LOWER(COALESCE(i.sucursal, '')) LIKE ?"
            params.append(f"%{sucursal.lower()}%")
        if modelo and modelo != "Todos":
            q += " AND i.modelo = ?"
            params.append(modelo)
        if ip and ip != "Todas":
            q += " AND m.ip = ?"
            params.append(ip)
        q += " ORDER BY m.fecha DESC, sucursal, m.ip"
        rows = conn.execute(q, params).fetchall()

    filas = []
    for r in rows:
        try:
            fecha_dt  = datetime.strptime(r[0], "%Y-%m-%d %H:%M:%S")
            fecha_str = fecha_dt.strftime("%d/%m/%Y %H:%M")
        except (ValueError, TypeError):
            fecha_str = str(r[0] or "")
        toner  = r[4]
        unidad = r[5]
        kit    = r[6]
        toner_str  = f"{toner*100:.1f}%"  if toner  is not None else "—"
        unidad_str = f"{unidad*100:.1f}%" if unidad is not None else "—"
        kit_str    = f"{kit*100:.1f}%"    if kit    is not None else "—"
        valores = [v for v in (toner, kit, unidad) if v is not None]
        tag     = clasificar_nivel(valores, umbral_bajo, umbral_medio)
        filas.append((fecha_str, r[1], r[2], r[3], toner_str, unidad_str, kit_str, tag))
    return filas


def db_cargar_ultimo_monitoreo(umbral_bajo, umbral_medio):
    """Retorna (filas, fecha_str) del monitoreo más reciente para poblar la tabla al inicio.
    filas: lista de tuplas (sucursal, ip, modelo, toner_str, unidad_str, kit_str, tag).
    """
    with db_connect() as conn:
        row_fecha = conn.execute("SELECT MAX(fecha) FROM monitoreos").fetchone()
        if not row_fecha or row_fecha[0] is None:
            return [], None
        ultima_fecha = row_fecha[0]
        rows = conn.execute("""
            SELECT COALESCE(i.sucursal, '') AS sucursal,
                   m.ip,
                   COALESCE(i.modelo, '')   AS modelo,
                   m.toner, m.unidad_imagen, m.kit_mantenimiento
            FROM monitoreos m
            LEFT JOIN impresoras i ON m.ip = i.ip
            WHERE m.fecha = ?
            ORDER BY sucursal, m.ip
        """, (ultima_fecha,)).fetchall()

    filas = []
    for r in rows:
        toner  = r[3]
        unidad = r[4]
        kit    = r[5]
        toner_str  = f"{toner*100:.1f}%"  if toner  is not None else "—"
        unidad_str = f"{unidad*100:.1f}%" if unidad is not None else "—"
        kit_str    = f"{kit*100:.1f}%"    if kit    is not None else "—"
        valores = [v for v in (toner, kit, unidad) if v is not None]
        tag     = clasificar_nivel(valores, umbral_bajo, umbral_medio)
        filas.append((r[0], r[1], r[2], fecha_ult, toner_str, unidad_str, kit_str, tag))
    return filas, ultima_fecha

# ---------------------------------------------------------------------------
# Lógica de negocio — monitoreo
# ---------------------------------------------------------------------------

def obtener_status(ip, modelo, evento_cancelar):
    """Consulta una impresora por HTTP. Retorna (toner, kit, unidad) como
    decimales 0-1, o (None, None, None) si no se puede obtener el dato."""
    if evento_cancelar.is_set():
        return None, None, None
    try:
        url = f"http://{ip}/cgi-bin/dynamic/printer/PrinterStatus.html"
        response = requests.get(url, timeout=5)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        porcentajes = re.findall(r'\b\d+%|\b\d+\.\d+%', soup.get_text())

        if modelo not in MODELOS_CONFIG:
            return None, None, None

        indices = MODELOS_CONFIG[modelo]

        def extraer(idx):
            if idx is None or len(porcentajes) <= idx:
                return None
            return float(porcentajes[idx].replace('%', '')) / 100

        return extraer(indices[0]), extraer(indices[1]), extraer(indices[2])
    except Exception as e:
        if not evento_cancelar.is_set():
            _log.error("IP %-15s  modelo %-22s  error: %s", ip, modelo, e)
        return None, None, None


def clasificar_nivel(valores, umbral_bajo, umbral_medio):
    """Clasifica consumibles según umbrales.
    Retorna: 'sin_datos', 'bajo', 'medio', o '' (nivel normal).
    """
    if not valores:
        return "sin_datos"
    m = min(valores)
    if m < umbral_bajo / 100:
        return "bajo"
    if m < umbral_medio / 100:
        return "medio"
    return ""


def predecir_agotamiento(fechas, valores):
    """Calcula la fecha estimada de agotamiento por regresión lineal.
    fechas: lista de datetime/date. valores: lista de floats 0-100 (NaN = sin dato).
    Retorna datetime o None si no se puede predecir.
    """
    pares = []
    for f, v in zip(fechas, valores):
        if f is None or v != v:          # v != v detecta NaN
            continue
        if not hasattr(f, 'hour'):
            f = datetime(f.year, f.month, f.day)
        pares.append((f, v))

    if len(pares) < 2:
        return None

    base = pares[0][0]
    x = np.array([(f - base).total_seconds() / 86400 for f, _ in pares], dtype=float)
    y = np.array([v for _, v in pares], dtype=float)

    coef = np.polyfit(x, y, 1)

    if coef[0] >= 0:
        return None

    dias = float(-coef[1] / coef[0])

    if dias > 3650:
        return None

    fecha_pred = base + timedelta(days=dias)

    if fecha_pred.date() <= datetime.today().date():
        return None

    return fecha_pred

# ---------------------------------------------------------------------------
# Ventana de stock de depósito
# ---------------------------------------------------------------------------

def abrir_stock_deposito():
    """Abre la ventana para gestionar el stock de insumos en depósito."""
    win = tk.Toplevel()

    win.title("Stock de Depósito")
    win.geometry("920x680")
    win.resizable(True, True)
    win.config(bg=BG_MAIN)

    modelos_catalogo = sorted({imp["modelo"] for imp in db_impresoras_todas() if imp["modelo"]})

    all_movs = []
    mov_page_size = 100
    mov_cur_page = 0

    # ── LabelFrame "Stock Actual" ────────────────────────────────────────────
    frame_stock = tk.LabelFrame(win, text="Stock Actual", bg=BG_MAIN,
                                font=FONT_BOLD, padx=8, pady=6)
    frame_stock.pack(fill="x", padx=10, pady=(10, 4))

    cols_stock   = ("Tipo Insumo", "Modelo Impresora", "En Depósito", "Mínimo", "Estado")
    col_w_stock  = (110, 240, 90, 70, 90)
    tree_stock   = Treeview(frame_stock, columns=cols_stock, show="headings", height=6)
    for col, w in zip(cols_stock, col_w_stock):
        tree_stock.heading(col, text=col)
        tree_stock.column(col, anchor="center", width=w)
    tree_stock.tag_configure("ok",      background="#C8E6C9", foreground="#1B5E20")
    tree_stock.tag_configure("bajo",    background="#FFF9C4", foreground="#E65100")
    tree_stock.tag_configure("critico", background="#FFCDD2", foreground="#B71C1C")
    sb_stock = Scrollbar(frame_stock, orient="vertical", command=tree_stock.yview)
    tree_stock.configure(yscrollcommand=sb_stock.set)
    tree_stock.pack(side="left", fill="both", expand=True)
    sb_stock.pack(side="right", fill="y")

    frame_stock_btns = tk.Frame(win, bg=BG_MAIN)
    frame_stock_btns.pack(fill="x", padx=10, pady=(0, 2))
    lbl_alertas = tk.Label(frame_stock_btns, text="", bg=BG_MAIN, font=("Segoe UI", 8),
                           fg="#B71C1C", anchor="w")
    lbl_alertas.pack(side="left", fill="x", expand=True)
    btn_editar = tk.Button(frame_stock_btns, text="Editar")
    btn_editar.pack(side="right", padx=(4, 0))
    btn_exportar_stock = tk.Button(frame_stock_btns, text="Exportar Excel")
    btn_exportar_stock.pack(side="right")
    for b in (btn_editar, btn_exportar_stock):
        _estilo_btn(b, primario=False)

    # ── LabelFrame "Registrar Entrada de Stock" ──────────────────────────────
    frame_entrada = tk.LabelFrame(win, text="Registrar Entrada de Stock", bg=BG_MAIN,
                                  font=FONT_BOLD, padx=8, pady=6)
    frame_entrada.pack(fill="x", padx=10, pady=(4, 4))

    lbl_kw = {"bg": BG_MAIN, "font": FONT_UI, "fg": "#555555"}

    tk.Label(frame_entrada, text="Tipo:", **lbl_kw).grid(row=0, column=0, sticky="e", padx=(0, 4))
    var_tipo_ent = tk.StringVar(value=TIPOS_INSUMO[0])
    combo_tipo_ent = Combobox(frame_entrada, textvariable=var_tipo_ent, values=TIPOS_INSUMO,
                              state="readonly", width=14)
    combo_tipo_ent.grid(row=0, column=1, sticky="w", padx=(0, 12))

    tk.Label(frame_entrada, text="Modelo:", **lbl_kw).grid(row=0, column=2, sticky="e", padx=(0, 4))
    var_modelo_ent = tk.StringVar()
    combo_modelo_ent = Combobox(frame_entrada, textvariable=var_modelo_ent,
                                values=modelos_catalogo, state="normal", width=22)
    combo_modelo_ent.grid(row=0, column=3, sticky="w", padx=(0, 12))

    tk.Label(frame_entrada, text="Cantidad:", **lbl_kw).grid(row=0, column=4, sticky="e", padx=(0, 4))
    spin_cant_ent = Spinbox(frame_entrada, from_=1, to=9999, width=6, font=FONT_UI)
    spin_cant_ent.grid(row=0, column=5, sticky="w", padx=(0, 12))

    tk.Label(frame_entrada, text="Observación:", **lbl_kw).grid(row=1, column=0, sticky="e",
                                                                 padx=(0, 4), pady=(6, 2))
    entry_obs = tk.Entry(frame_entrada, width=40, font=FONT_UI)
    entry_obs.grid(row=1, column=1, columnspan=4, sticky="w", pady=(6, 2))

    btn_agregar = tk.Button(frame_entrada, text="Agregar al Depósito")
    _estilo_btn(btn_agregar, primario=True)
    btn_agregar.grid(row=1, column=5, padx=(4, 0), pady=(6, 2))

    # ── LabelFrame "Historial de Movimientos" ────────────────────────────────
    frame_hist = tk.LabelFrame(win, text="Historial de Movimientos", bg=BG_MAIN,
                               font=FONT_BOLD, padx=8, pady=6)
    frame_hist.pack(fill="both", expand=True, padx=10, pady=(4, 4))

    frame_filtros_h = tk.Frame(frame_hist, bg=BG_MAIN)
    frame_filtros_h.pack(fill="x", pady=(0, 6))

    tk.Label(frame_filtros_h, text="Tipo Insumo:", **lbl_kw).pack(side="left", padx=(0, 4))
    var_tipo_h = tk.StringVar(value="Todos")
    Combobox(frame_filtros_h, textvariable=var_tipo_h,
             values=["Todos"] + TIPOS_INSUMO, state="readonly", width=13).pack(side="left", padx=(0, 8))

    tk.Label(frame_filtros_h, text="Movimiento:", **lbl_kw).pack(side="left", padx=(0, 4))
    var_mov_tipo = tk.StringVar(value="Todos")
    Combobox(frame_filtros_h, textvariable=var_mov_tipo,
             values=["Todos", "Entrada", "Salida", "Ajuste"],
             state="readonly", width=10).pack(side="left", padx=(0, 8))

    tk.Label(frame_filtros_h, text="Modelo:", **lbl_kw).pack(side="left", padx=(0, 4))
    var_modelo_h = tk.StringVar(value="Todos")
    mod_h_list = ["Todos"] + modelos_catalogo
    Combobox(frame_filtros_h, textvariable=var_modelo_h,
             values=mod_h_list, state="readonly", width=18).pack(side="left", padx=(0, 8))

    tk.Label(frame_filtros_h, text="Desde:", **lbl_kw).pack(side="left", padx=(0, 4))
    entry_desde_h = tk.Entry(frame_filtros_h, width=10, font=FONT_UI)
    entry_desde_h.pack(side="left", padx=(0, 6))

    tk.Label(frame_filtros_h, text="Hasta:", **lbl_kw).pack(side="left", padx=(0, 4))
    entry_hasta_h = tk.Entry(frame_filtros_h, width=10, font=FONT_UI)
    entry_hasta_h.pack(side="left", padx=(0, 6))

    tk.Label(frame_filtros_h, text="(DD/MM/YYYY)", bg=BG_MAIN,
             font=("Segoe UI", 7), fg="#AAAAAA").pack(side="left", padx=(0, 6))

    btn_aplicar_h = tk.Button(frame_filtros_h, text="Aplicar")
    _estilo_btn(btn_aplicar_h, primario=True)
    btn_aplicar_h.pack(side="left", padx=(0, 4))
    btn_todo_h = tk.Button(frame_filtros_h, text="Todo")
    _estilo_btn(btn_todo_h, primario=False)
    btn_todo_h.pack(side="left")

    cols_hist  = ("Fecha", "Movimiento", "Tipo Insumo", "Modelo", "Cantidad", "Observación")
    col_w_hist = (130, 75, 105, 200, 70, 220)
    frame_tree_h = tk.Frame(frame_hist, bg=BG_MAIN)
    frame_tree_h.pack(fill="both", expand=True)

    tree_hist = Treeview(frame_tree_h, columns=cols_hist, show="headings", height=8)
    for col, w in zip(cols_hist, col_w_hist):
        tree_hist.heading(col, text=col)
        tree_hist.column(col, anchor="center" if col != "Observación" else "w", width=w)
    tree_hist.tag_configure("entrada", foreground="#1B5E20")
    tree_hist.tag_configure("salida",  foreground="#B71C1C")
    tree_hist.tag_configure("ajuste",  foreground="#1565C0")
    sb_hist = Scrollbar(frame_tree_h, orient="vertical", command=tree_hist.yview)
    tree_hist.configure(yscrollcommand=sb_hist.set)
    tree_hist.pack(side="left", fill="both", expand=True)
    sb_hist.pack(side="right", fill="y")

    # ── Barra inferior movimientos ───────────────────────────────────────────
    frame_mov_pie = tk.Frame(win, bg=BG_MAIN)
    frame_mov_pie.pack(fill="x", padx=10, pady=(0, 8))

    btn_prev_mov = tk.Button(frame_mov_pie, text="<<", width=3)
    btn_prev_mov.pack(side="left", padx=(0, 2))
    lbl_mov_pag = tk.Label(frame_mov_pie, text="", width=16, anchor="center", fg="gray")
    lbl_mov_pag.pack(side="left")
    btn_next_mov = tk.Button(frame_mov_pie, text=">>", width=3)
    btn_next_mov.pack(side="left", padx=(2, 0))

    btn_eliminar_mov = tk.Button(frame_mov_pie, text="Eliminar movimiento")
    btn_eliminar_mov.pack(side="left", padx=(10, 0))
    btn_exportar_mov = tk.Button(frame_mov_pie, text="Exportar Excel")
    btn_exportar_mov.pack(side="right")
    for b in (btn_prev_mov, btn_next_mov, btn_eliminar_mov, btn_exportar_mov):
        _estilo_btn(b, primario=False)

    # ── Funciones internas ───────────────────────────────────────────────────

    def _cargar_stock():
        tree_stock.delete(*tree_stock.get_children())
        stock = db_stock_obtener()
        criticos = bajos = 0
        for r in stock:
            cant = r["cantidad"]
            mini = r["stock_minimo"]
            if cant <= mini:
                tag = "critico"; estado = "Crítico"; criticos += 1
            elif cant <= mini * 2:
                tag = "bajo"; estado = "Bajo"; bajos += 1
            else:
                tag = "ok"; estado = "OK"
            tree_stock.insert("", "end",
                              values=(r["tipo_insumo"], r["modelo_impresora"],
                                      cant, mini, estado), tags=(tag,))
        partes = []
        if criticos: partes.append(f"\u26a0 {criticos} modelo(s) con stock crítico")
        if bajos:    partes.append(f"\u26a0 {bajos} modelo(s) en nivel bajo")
        lbl_alertas.config(text="  ".join(partes) if partes else "",
                           fg="#B71C1C" if criticos else "#E65100")

    def _editar_fila():
        sel = tree_stock.selection()
        if not sel:
            messagebox.showwarning("Sin selección", "Seleccione un modelo.", parent=win)
            return
        vals = tree_stock.item(sel[0], "values")
        tipo_sel, modelo_sel = vals[0], vals[1]
        cant_sel, min_sel = int(vals[2]), int(vals[3])
        popup = tk.Toplevel(win)
        popup.title("Editar stock"); popup.geometry("300x170"); popup.resizable(False, False)
        popup.config(bg=BG_MAIN); popup.grab_set()
        tk.Label(popup, text=f"{tipo_sel}\n{modelo_sel}",
                 bg=BG_MAIN, font=FONT_BOLD, wraplength=280).pack(pady=(10, 6))
        fp = tk.Frame(popup, bg=BG_MAIN); fp.pack()
        tk.Label(fp, text="Cantidad en depósito:", bg=BG_MAIN, font=FONT_UI,
                 width=20, anchor="e").grid(row=0, column=0, padx=(0,6), pady=3)
        sc = Spinbox(fp, from_=0, to=9999, width=7, font=FONT_UI)
        sc.delete(0, tk.END); sc.insert(0, str(cant_sel))
        sc.grid(row=0, column=1, sticky="w")
        tk.Label(fp, text="Stock mínimo:", bg=BG_MAIN, font=FONT_UI,
                 width=20, anchor="e").grid(row=1, column=0, padx=(0,6), pady=3)
        sm = Spinbox(fp, from_=0, to=999, width=7, font=FONT_UI)
        sm.delete(0, tk.END); sm.insert(0, str(min_sel))
        sm.grid(row=1, column=1, sticky="w")
        def _guardar():
            try:
                nc = int(sc.get()); nm = int(sm.get())
            except ValueError: return
            with db_connect() as conn:
                conn.execute(
                    "UPDATE stock_deposito SET cantidad=?, stock_minimo=? "
                    "WHERE tipo_insumo=? AND modelo_impresora=?",
                    (nc, nm, tipo_sel, modelo_sel))
                if nc != cant_sel:
                    delta = nc - cant_sel; signo = "+ " if delta > 0 else "- "
                    conn.execute(
                        "INSERT INTO movimientos_stock "
                        "(fecha, tipo, tipo_insumo, modelo_impresora, cantidad, observacion) "
                        "VALUES (?, 'ajuste', ?, ?, ?, ?)",
                        (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                         tipo_sel, modelo_sel, abs(delta),
                         f"Ajuste manual ({signo}{abs(delta)})"))
            _cargar_stock(); _cargar_historial(); popup.destroy()
        fbp = tk.Frame(popup, bg=BG_MAIN); fbp.pack(pady=8)
        bg = tk.Button(fbp, text="Guardar", command=_guardar)
        _estilo_btn(bg, primario=True); bg.pack(side="left", padx=4)
        bc = tk.Button(fbp, text="Cancelar", command=popup.destroy)
        _estilo_btn(bc, primario=False); bc.pack(side="left", padx=4)

    def _exportar_stock():
        items = tree_stock.get_children()
        if not items:
            messagebox.showwarning("Sin datos", "No hay stock para exportar.", parent=win)
            return
        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"stock_{datetime.now().strftime('%Y%m%d')}.xlsx", parent=win)
        if not ruta: return
        wb = Workbook(); ws = wb.active; ws.title = "Stock"
        headers = list(cols_stock); widths = [14, 26, 12, 8, 8]
        hf = PatternFill("solid", fgColor="4472C4"); hfont = Font(bold=True, color="FFFFFF")
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=col, value=h); c.fill = hf; c.font = hfont
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[c.column_letter].width = w
        fills = {"critico": PatternFill("solid", fgColor="FFCDD2"),
                 "bajo":    PatternFill("solid", fgColor="FFF9C4")}
        for item in items:
            ws.append(list(tree_stock.item(item, "values")))
            tag = (tree_stock.item(item, "tags") or ("",))[0]
            if tag in fills:
                for col in range(1, len(headers)+1):
                    ws.cell(row=ws.max_row, column=col).fill = fills[tag]
        wb.save(ruta)
        messagebox.showinfo("Exportado", f"Stock guardado:\n{ruta}", parent=win)

    def _registrar_entrada():
        tipo, modelo = var_tipo_ent.get().strip(), var_modelo_ent.get().strip()
        obs = entry_obs.get().strip()
        if not tipo or not modelo:
            messagebox.showwarning("Campos requeridos", "Seleccione tipo y modelo.", parent=win)
            return
        try:
            cant = int(spin_cant_ent.get())
            if cant <= 0: raise ValueError
        except ValueError:
            messagebox.showwarning("Cantidad inválida", "Debe ser un entero positivo.", parent=win)
            return
        db_stock_agregar_entrada(tipo, modelo, cant, obs)
        entry_obs.delete(0, tk.END); _cargar_stock(); _cargar_historial()

    def _cargar_historial():
        nonlocal all_movs, mov_cur_page
        mov_cur_page = 0
        for i, col in enumerate(cols_hist):
            tree_hist.heading(col, text=col)
        rows = db_movimientos_stock()
        tipo_filtro      = var_tipo_h.get()
        modelo_filtro    = var_modelo_h.get()
        mov_tipo_filtro  = var_mov_tipo.get()
        tipo_map = {"Entrada": "entrada", "Salida": "salida", "Ajuste": "ajuste"}
        filtrados = []
        for r in rows:
            if tipo_filtro != "Todos" and r[2] != tipo_filtro: continue
            if modelo_filtro != "Todos" and r[3] != modelo_filtro: continue
            if mov_tipo_filtro != "Todos" and r[1] != tipo_map.get(mov_tipo_filtro, ""): continue
            filtrados.append(r)
        all_movs = filtrados
        _mostrar_mov_pagina()

    def _mostrar_mov_pagina():
        inicio = mov_cur_page * mov_page_size
        pagina = all_movs[inicio:inicio + mov_page_size]
        tree_hist.delete(*tree_hist.get_children())
        for r in pagina:
            try:
                fecha_str = datetime.strptime(r[0], "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M")
            except (ValueError, TypeError):
                fecha_str = str(r[0] or "")
            tipo_label = {"entrada": "Entrada", "salida": "Salida", "ajuste": "Ajuste"}.get(r[1], r[1].capitalize())
            tag_mov = r[1] if r[1] in ("entrada", "salida", "ajuste") else "entrada"
            tree_hist.insert("", "end",
                             values=(fecha_str, tipo_label, r[2], r[3], r[4], r[5] or ""),
                             tags=(tag_mov,))
        total = len(all_movs)
        pag_tot = max(1, (total + mov_page_size - 1) // mov_page_size)
        lbl_mov_pag.config(text=f"Pág. {mov_cur_page + 1} / {pag_tot}")
        btn_prev_mov.config(state="normal" if mov_cur_page > 0 else "disabled")
        btn_next_mov.config(state="normal" if inicio + mov_page_size < total else "disabled")

    def _todo_historial():
        var_tipo_h.set("Todos"); var_modelo_h.set("Todos"); var_mov_tipo.set("Todos")
        entry_desde_h.delete(0, tk.END); entry_hasta_h.delete(0, tk.END)
        _cargar_historial()

    def _eliminar_mov():
        sel = tree_hist.selection()
        if not sel:
            messagebox.showinfo("Sin selección", "Seleccione un movimiento.", parent=win)
            return
        vals = tree_hist.item(sel[0], "values")
        if not messagebox.askyesno("Confirmar",
                                   f"¿Eliminar movimiento del {vals[0]}?\n{vals[2]} {vals[3]} x{vals[4]}",
                                   parent=win):
            return
        try:
            idx = list(tree_hist.get_children()).index(sel[0])
            global_idx = mov_cur_page * mov_page_size + idx
            if global_idx < len(all_movs):
                r = all_movs[global_idx]
                with db_connect() as conn:
                    conn.execute("""
                        DELETE FROM movimientos_stock WHERE rowid IN (
                            SELECT rowid FROM movimientos_stock
                            WHERE fecha=? AND tipo=? AND tipo_insumo=? AND modelo_impresora=? AND cantidad=?
                            LIMIT 1
                        )
                    """, (r[0], r[1], r[2], r[3], r[4]))
                _cargar_historial()
        except (ValueError, IndexError):
            pass

    def _exportar_mov():
        if not all_movs:
            messagebox.showwarning("Sin datos", "No hay movimientos para exportar.", parent=win)
            return
        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"movimientos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", parent=win)
        if not ruta: return
        wb = Workbook(); ws = wb.active; ws.title = "Movimientos"
        headers = list(cols_hist); widths = [17, 10, 14, 24, 8, 30]
        hf = PatternFill("solid", fgColor="CCCCCC"); hfont = Font(bold=True)
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=col, value=h); c.fill = hf; c.font = hfont
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[c.column_letter].width = w
        fills = {"entrada": PatternFill("solid", fgColor="C8E6C9"),
                 "salida":  PatternFill("solid", fgColor="FFCDD2"),
                 "ajuste":  PatternFill("solid", fgColor="BBDEFB")}
        for r in all_movs:
            try:
                fecha_str = datetime.strptime(r[0], "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M")
            except (ValueError, TypeError):
                fecha_str = str(r[0] or "")
            tipo_label = {"entrada": "Entrada", "salida": "Salida", "ajuste": "Ajuste"}.get(r[1], r[1].capitalize())
            ws.append([fecha_str, tipo_label, r[2], r[3], r[4], r[5] or ""])
            if r[1] in fills:
                for col in range(1, len(headers)+1):
                    ws.cell(row=ws.max_row, column=col).fill = fills[r[1]]
        wb.save(ruta)
        messagebox.showinfo("Exportado", f"Movimientos guardados:\n{ruta}", parent=win)

    def _pag_mov(delta):
        nonlocal mov_cur_page
        nueva = mov_cur_page + delta
        total_pag = max(1, (len(all_movs) + mov_page_size - 1) // mov_page_size)
        if 0 <= nueva < total_pag:
            mov_cur_page = nueva
            _mostrar_mov_pagina()

    # ── Comandos ─────────────────────────────────────────────────────────────
    for i, col in enumerate(cols_stock):
        tree_stock.heading(col, text=col)
    for i, col in enumerate(cols_hist):
        tree_hist.heading(col, text=col)

    btn_editar.config(command=_editar_fila)
    btn_exportar_stock.config(command=_exportar_stock)
    btn_agregar.config(command=_registrar_entrada)
    btn_aplicar_h.config(command=_cargar_historial)
    btn_todo_h.config(command=_todo_historial)
    btn_prev_mov.config(command=lambda: _pag_mov(-1))
    btn_next_mov.config(command=lambda: _pag_mov(1))
    btn_eliminar_mov.config(command=_eliminar_mov)
    btn_exportar_mov.config(command=_exportar_mov)

    # ── Ordenamiento por columna ─────────────────────────────────────────────
    def _ordenar_stock(col_idx):
        items = [(tree_stock.set(iid, col_idx), iid) for iid in tree_stock.get_children("")]
        try:
            items.sort(key=lambda x: (x[0].isdigit(), int(x[0])) if x[0].isdigit() else (False, x[0].lower()))
        except (ValueError, TypeError):
            items.sort(key=lambda x: str(x[0]).lower())
        for idx, (_, iid) in enumerate(items):
            tree_stock.move(iid, "", idx)

    def _ordenar_mov(col_idx):
        nonlocal all_movs, mov_cur_page
        mov_cur_page = 0
        def _mk(r, ci):
            v = r[ci]
            if ci == 0:
                try: return datetime.strptime(v, "%Y-%m-%d %H:%M:%S")
                except: return datetime.min
            if ci == 4:
                try: return int(v)
                except: return 0
            return v.lower()
        all_movs.sort(key=lambda r: _mk(r, col_idx))
        _mostrar_mov_pagina()

    for i, col in enumerate(cols_stock):
        tree_stock.heading(col, text=col, command=lambda i=i: _ordenar_stock(i))
    for i, col in enumerate(cols_hist):
        tree_hist.heading(col, text=col, command=lambda i=i: _ordenar_mov(i))

    # ── Doble clic en stock → editar ────────────────────────────────────────
    tree_stock.bind("<<TreeviewDoubleClick>>", lambda _: _editar_fila())

    # ── Filtro automático movimientos ────────────────────────────────────────
    var_tipo_h.trace_add("write", lambda *_: _cargar_historial())
    var_modelo_h.trace_add("write", lambda *_: _cargar_historial())
    var_mov_tipo.trace_add("write", lambda *_: _cargar_historial())

    # ── Autocompletar modelo según tipo de insumo ────────────────────────────
    def _filtrar_modelos_por_tipo(*_):
        tipo = var_tipo_ent.get().strip()
        if not tipo:
            combo_modelo_ent.config(values=modelos_catalogo)
            return
        with db_connect() as conn:
            rows = conn.execute(
                "SELECT DISTINCT modelo_impresora FROM stock_deposito WHERE tipo_insumo=? ORDER BY modelo_impresora",
                (tipo,)).fetchall()
        modelos_filtrados = [r[0] for r in rows] if rows else []
        if not modelos_filtrados:
            modelos_filtrados = modelos_catalogo
        combo_modelo_ent.config(values=modelos_filtrados)
        if var_modelo_ent.get() not in modelos_filtrados:
            var_modelo_ent.set("")

    var_tipo_ent.trace_add("write", _filtrar_modelos_por_tipo)

    # ── Carga inicial ─────────────────────────────────────────────────────────
    _cargar_stock()
    _cargar_historial()


# ---------------------------------------------------------------------------
# Anulación y edición de envíos
# ---------------------------------------------------------------------------

def db_anular_envio(envio_id):
    """Marca un envío como anulado y restaura el stock."""
    with db_connect() as conn:
        envio = conn.execute(
            "SELECT tipo_insumo, modelo_impresora, cantidad FROM envios WHERE id=? AND anulado=0",
            (envio_id,)).fetchone()
        if not envio:
            return False
        tipo, modelo, cantidad = envio
        conn.execute("UPDATE envios SET anulado=1 WHERE id=?", (envio_id,))
        conn.execute(
            "INSERT OR IGNORE INTO stock_deposito (tipo_insumo, modelo_impresora, cantidad) "
            "VALUES (?, ?, 0)", (tipo, modelo))
        conn.execute(
            "UPDATE stock_deposito SET cantidad = cantidad + ? WHERE tipo_insumo=? AND modelo_impresora=?",
            (cantidad, tipo, modelo))
        conn.execute(
            "INSERT INTO movimientos_stock "
            "(fecha, tipo, tipo_insumo, modelo_impresora, cantidad, observacion, envio_id) "
            "VALUES (?, 'entrada', ?, ?, ?, ?, ?)",
            (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), tipo, modelo, cantidad,
             f"Devolución por anulación de envío #{envio_id}", envio_id))
        return True


def db_editar_envio(envio_id, nueva_cantidad):
    """Cambia la cantidad de un envío y ajusta el stock."""
    with db_connect() as conn:
        envio = conn.execute(
            "SELECT tipo_insumo, modelo_impresora, cantidad FROM envios WHERE id=? AND anulado=0",
            (envio_id,)).fetchone()
        if not envio:
            return False
        tipo, modelo, cantidad_anterior = envio
        if nueva_cantidad == cantidad_anterior:
            return True
        conn.execute("UPDATE envios SET cantidad=? WHERE id=?", (nueva_cantidad, envio_id))
        conn.execute(
            "INSERT OR IGNORE INTO stock_deposito (tipo_insumo, modelo_impresora, cantidad) "
            "VALUES (?, ?, 0)", (tipo, modelo))
        if nueva_cantidad > cantidad_anterior:
            delta = nueva_cantidad - cantidad_anterior
            conn.execute(
                "UPDATE stock_deposito SET cantidad = cantidad - ? WHERE tipo_insumo=? AND modelo_impresora=?",
                (delta, tipo, modelo))
            obs = f"Ajuste por edición de envío #{envio_id} ({cantidad_anterior}→{nueva_cantidad})"
            conn.execute(
                "INSERT INTO movimientos_stock "
                "(fecha, tipo, tipo_insumo, modelo_impresora, cantidad, observacion, envio_id) "
                "VALUES (?, 'salida', ?, ?, ?, ?, ?)",
                (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), tipo, modelo, delta, obs, envio_id))
        else:
            delta = cantidad_anterior - nueva_cantidad
            conn.execute(
                "UPDATE stock_deposito SET cantidad = cantidad + ? WHERE tipo_insumo=? AND modelo_impresora=?",
                (delta, tipo, modelo))
            obs = f"Ajuste por edición de envío #{envio_id} ({cantidad_anterior}→{nueva_cantidad})"
            conn.execute(
                "INSERT INTO movimientos_stock "
                "(fecha, tipo, tipo_insumo, modelo_impresora, cantidad, observacion, envio_id) "
                "VALUES (?, 'entrada', ?, ?, ?, ?, ?)",
                (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), tipo, modelo, delta, obs, envio_id))
        return True


# ---------------------------------------------------------------------------
# Ventana de envío de insumos
# ---------------------------------------------------------------------------

def abrir_envio_insumos():
    """Abre la ventana de envío de insumos con registro, anulación y edición."""
    win = tk.Toplevel()
    win.title("Envío de Insumos")
    win.geometry("860x620")
    win.resizable(True, True)

    # Construir mapa sucursal → lista de IPs, e ip → datos de impresora
    impresoras_db  = db_impresoras_todas(activas_solo=True)
    sucursales_ips = {}
    ip_info        = {}
    for _imp in impresoras_db:
        suc = _imp["sucursal"]
        if suc:
            sucursales_ips.setdefault(suc, []).append(_imp["ip"])
        ip_info[_imp["ip"]] = _imp
    sucursales_lista = sorted(sucursales_ips.keys())

    anio_actual   = datetime.today().year
    anios         = ["Todos"] + [str(anio_actual - 2 + i) for i in range(5)]
    meses         = ["Todos", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                     "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    # Mapeo iid → id de DB
    id_map = {}

    # ------------------------------------------------------------------
    # Sección: Registrar Envío
    # ------------------------------------------------------------------
    frame_form = tk.LabelFrame(win, text="Registrar Envío", padx=8, pady=6)
    frame_form.pack(fill="x", padx=10, pady=(10, 4))

    tk.Label(frame_form, text="Fecha:").grid(row=0, column=0, sticky="e", padx=(0, 4))
    var_fecha = tk.StringVar(value=datetime.today().strftime("%d/%m/%Y"))
    tk.Entry(frame_form, textvariable=var_fecha, width=12).grid(row=0, column=1, sticky="w", padx=(0, 14))

    tk.Label(frame_form, text="Sucursal:").grid(row=0, column=2, sticky="e", padx=(0, 4))
    var_sucursal   = tk.StringVar()
    combo_sucursal = Combobox(frame_form, textvariable=var_sucursal, values=sucursales_lista,
                              state="readonly", width=19)
    combo_sucursal.grid(row=0, column=3, sticky="w", padx=(0, 14))

    tk.Label(frame_form, text="IP:").grid(row=0, column=4, sticky="e", padx=(0, 4))
    var_ip_envio   = tk.StringVar()
    combo_ip_envio = Combobox(frame_form, textvariable=var_ip_envio, values=[],
                              state="readonly", width=13)
    combo_ip_envio.grid(row=0, column=5, sticky="w")

    # Fila 1: Info de la impresora seleccionada (solo lectura)
    var_info_modelo = tk.StringVar(value="—")
    var_info_nombre = tk.StringVar(value="—")
    var_info_sn     = tk.StringVar(value="—")
    frame_info = tk.Frame(frame_form, bg="#f0f0f0", bd=1, relief="groove")
    frame_info.grid(row=1, column=0, columnspan=6, sticky="ew", padx=4, pady=(6, 2))
    tk.Label(frame_info, text="Modelo:",  font=("", 8, "bold"), bg="#f0f0f0").pack(side="left", padx=(8, 2), pady=3)
    tk.Label(frame_info, textvariable=var_info_modelo, bg="#f0f0f0", width=18, anchor="w", fg="#333333").pack(side="left", padx=(0, 14))
    tk.Label(frame_info, text="Nombre:",  font=("", 8, "bold"), bg="#f0f0f0").pack(side="left", padx=(0, 2))
    tk.Label(frame_info, textvariable=var_info_nombre, bg="#f0f0f0", width=18, anchor="w", fg="#333333").pack(side="left", padx=(0, 14))
    tk.Label(frame_info, text="N° Serie:", font=("", 8, "bold"), bg="#f0f0f0").pack(side="left", padx=(0, 2))
    tk.Label(frame_info, textvariable=var_info_sn,     bg="#f0f0f0", width=14, anchor="w", fg="#333333").pack(side="left")

    def _limpiar_info():
        var_info_modelo.set("—")
        var_info_nombre.set("—")
        var_info_sn.set("—")

    def al_cambiar_sucursal(_=None):
        suc = var_sucursal.get()
        combo_ip_envio.config(values=[""] + sucursales_ips.get(suc, []))
        var_ip_envio.set("")
        _limpiar_info()

    combo_sucursal.bind("<<ComboboxSelected>>", al_cambiar_sucursal)

    def al_cambiar_ip(_=None):
        ip = var_ip_envio.get()
        if ip and ip in ip_info:
            imp = ip_info[ip]
            var_info_modelo.set(imp["modelo"] or "—")
            var_info_nombre.set(imp["nombre"] or "—")
            var_info_sn.set(imp.get("sn", "") or "—")
        else:
            _limpiar_info()

    combo_ip_envio.bind("<<ComboboxSelected>>", al_cambiar_ip)

    # Fila 2: Tipo · Cantidad · Botón Registrar
    tk.Label(frame_form, text="Tipo:").grid(row=2, column=0, sticky="e", padx=(0, 4), pady=(6, 2))
    var_tipo = tk.StringVar(value=TIPOS_INSUMO[0])
    Combobox(frame_form, textvariable=var_tipo, values=TIPOS_INSUMO,
             state="readonly", width=14).grid(row=2, column=1, sticky="w", padx=(0, 14), pady=(6, 2))

    tk.Label(frame_form, text="Cantidad:").grid(row=2, column=2, sticky="e", padx=(0, 4), pady=(6, 2))
    spin_cantidad = Spinbox(frame_form, from_=1, to=999, width=5)
    spin_cantidad.grid(row=2, column=3, sticky="w", pady=(6, 2))

    btn_registrar = tk.Button(frame_form, text="  Registrar Envío  ",
                              bg="#4CAF50", fg="white", font=("", 9, "bold"))
    btn_registrar.grid(row=2, column=4, columnspan=2, padx=(14, 0), pady=(6, 2))

    # ------------------------------------------------------------------
    # Filtros
    # ------------------------------------------------------------------
    frame_filtro = tk.Frame(win)
    frame_filtro.pack(fill="x", padx=10, pady=(4, 4))

    tk.Label(frame_filtro, text="Sucursal:").pack(side="left", padx=(0, 4))
    var_filtro_suc = tk.StringVar()
    tk.Entry(frame_filtro, textvariable=var_filtro_suc, width=16).pack(side="left", padx=(0, 8))

    tk.Label(frame_filtro, text="Año:").pack(side="left", padx=(0, 4))
    var_anio = tk.StringVar(value="Todos")
    Combobox(frame_filtro, textvariable=var_anio, values=anios,
             state="readonly", width=7).pack(side="left", padx=(0, 6))

    tk.Label(frame_filtro, text="Mes:").pack(side="left", padx=(0, 4))
    var_mes = tk.StringVar(value="Todos")
    Combobox(frame_filtro, textvariable=var_mes, values=meses,
             state="readonly", width=11).pack(side="left", padx=(0, 6))

    tk.Label(frame_filtro, text="Estado:").pack(side="left", padx=(0, 4))
    var_filtro_estado = tk.StringVar(value="Activos")
    Combobox(frame_filtro, textvariable=var_filtro_estado,
             values=["Todos", "Activos", "Anulados"],
             state="readonly", width=8).pack(side="left", padx=(0, 8))

    btn_filtrar = tk.Button(frame_filtro, text="Filtrar")
    btn_filtrar.pack(side="left", padx=(0, 4))
    btn_todos = tk.Button(frame_filtro, text="Ver todos")
    btn_todos.pack(side="left")

    # ------------------------------------------------------------------
    # Tabla de envíos
    # ------------------------------------------------------------------
    cols_env   = ("Fecha", "Sucursal", "IP", "Tipo Insumo", "Modelo Impresora", "Cant.", "Estado")
    col_widths = (90, 160, 110, 100, 175, 50, 75)

    frame_tree = tk.Frame(win)
    frame_tree.pack(fill="both", expand=True, padx=10, pady=(4, 0))

    tree_env = Treeview(frame_tree, columns=cols_env, show="headings", height=10)
    for col, w in zip(cols_env, col_widths):
        tree_env.heading(col, text=col)
        tree_env.column(col, anchor="center", width=w)
    tree_env.tag_configure("anulado", foreground="#999999")
    sb_env = Scrollbar(frame_tree, orient="vertical", command=tree_env.yview)
    tree_env.configure(yscroll=sb_env.set)
    tree_env.pack(side="left", fill="both", expand=True)
    sb_env.pack(side="right", fill="y")

    # ------------------------------------------------------------------
    # Botones de acción (anular / editar)
    # ------------------------------------------------------------------
    frame_acciones = tk.Frame(win)
    frame_acciones.pack(fill="x", padx=10, pady=(4, 2))
    btn_anular = tk.Button(frame_acciones, text="Anular Envío", state="disabled")
    btn_anular.pack(side="left", padx=(0, 6))
    btn_editar = tk.Button(frame_acciones, text="Editar Cantidad", state="disabled")
    btn_editar.pack(side="left")

    lbl_resumen = tk.Label(win, text="", anchor="w", font=("", 9))
    lbl_resumen.pack(fill="x", padx=12, pady=(2, 8))

    # ------------------------------------------------------------------
    # Funciones internas
    # ------------------------------------------------------------------
    def actualizar_tabla(filas):
        tree_env.delete(*tree_env.get_children())
        id_map.clear()
        for f in filas:
            envio_id, fecha_str, sucursal, ip, tipo, modelo, cantidad, anulado = f
            estado_str = "Anulado" if anulado else "Activo"
            tags = ("anulado",) if anulado else ()
            iid = tree_env.insert("", "end",
                                  values=(fecha_str, sucursal, ip, tipo, modelo, cantidad, estado_str),
                                  tags=tags)
            id_map[iid] = envio_id
        toner_total  = sum(f[6] for f in filas if f[4] == "Tóner" and not f[7] and isinstance(f[6], (int, float)))
        unidad_total = sum(f[6] for f in filas if f[4] == "Unidad Imagen" and not f[7] and isinstance(f[6], (int, float)))
        partes = []
        if toner_total:  partes.append(f"Tóner: {toner_total} ud.")
        if unidad_total: partes.append(f"Unidad Imagen: {unidad_total} ud.")
        total = len(filas)
        anulados = sum(1 for f in filas if f[7])
        if partes:
            lbl_resumen.config(text=f"{total} envío(s) — {anulados} anulado(s)  |  " + "   |   ".join(partes), fg="black")
        else:
            lbl_resumen.config(text=f"{total} envío(s) — {anulados} anulado(s)", fg="gray")

    def filtrar():
        suc        = var_filtro_suc.get().strip()
        anio_str   = var_anio.get()
        anio       = int(anio_str) if anio_str != "Todos" else None
        mes        = meses.index(var_mes.get())
        estado     = var_filtro_estado.get()
        filtro_an  = {"Todos": None, "Activos": False, "Anulados": True}.get(estado, None)
        actualizar_tabla(db_cargar_envios(sucursal_filtro=suc, anio=anio, mes=mes, filtro_anulado=filtro_an))

    def mostrar_todos():
        var_filtro_suc.set("")
        var_anio.set("Todos")
        var_mes.set("Todos")
        var_filtro_estado.set("Activos")
        actualizar_tabla(db_cargar_envios(filtro_anulado=False))

    def registrar():
        fecha_str = var_fecha.get().strip()
        try:
            fecha_dt = datetime.strptime(fecha_str, "%d/%m/%Y")
        except ValueError:
            messagebox.showwarning("Fecha inválida", "Use el formato DD/MM/YYYY.", parent=win)
            return

        sucursal = var_sucursal.get().strip()
        if not sucursal:
            messagebox.showwarning("Campo requerido", "Seleccione una sucursal.", parent=win)
            return

        ip = var_ip_envio.get().strip()
        if not ip:
            messagebox.showwarning("Campo requerido", "Seleccione una impresora (IP).", parent=win)
            return

        modelo = ip_info.get(ip, {}).get("modelo", "")

        try:
            cantidad = int(spin_cantidad.get())
        except ValueError:
            messagebox.showwarning("Cantidad inválida", "La cantidad debe ser un número entero.", parent=win)
            return

        try:
            db_registrar_envio(fecha_dt, sucursal, ip, var_tipo.get(), modelo, cantidad)
            messagebox.showinfo("Registrado", "Envío registrado correctamente.", parent=win)
            filtrar()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar:\n{e}", parent=win)

    def anular_envio():
        sel = tree_env.selection()
        if not sel:
            return
        iid = sel[0]
        envio_id = id_map.get(iid)
        if envio_id is None:
            return
        vals = tree_env.item(iid, "values")
        if vals[6] == "Anulado":
            messagebox.showinfo("Ya anulado", "Este envío ya fue anulado.", parent=win)
            return
        if not messagebox.askyesno("Confirmar anulación",
                                   f"¿Anular envío a {vals[1]} ({vals[3]}, {vals[4]} x{vals[5]})?\n\n"
                                   "El stock se restaurará automáticamente.",
                                   parent=win):
            return
        if db_anular_envio(envio_id):
            messagebox.showinfo("Anulado", "Envío anulado y stock restaurado.", parent=win)
            filtrar()
        else:
            messagebox.showerror("Error", "No se pudo anular el envío.", parent=win)

    def editar_envio():
        sel = tree_env.selection()
        if not sel:
            return
        iid = sel[0]
        envio_id = id_map.get(iid)
        if envio_id is None:
            return
        vals = tree_env.item(iid, "values")
        if vals[6] == "Anulado":
            messagebox.showinfo("Anulado", "No se puede editar un envío anulado.", parent=win)
            return

        popup = tk.Toplevel(win)
        popup.title("Editar cantidad")
        popup.geometry("320x150")
        popup.resizable(False, False)
        popup.config(bg=BG_MAIN)
        popup.grab_set()

        tk.Label(popup, text=f"{vals[1]} — {vals[3]} ({vals[4]})",
                 bg=BG_MAIN, font=FONT_BOLD, wraplength=280).pack(pady=(10, 6))

        frame_p = tk.Frame(popup, bg=BG_MAIN)
        frame_p.pack()
        tk.Label(frame_p, text="Cantidad:", bg=BG_MAIN, font=FONT_UI).pack(side="left", padx=(0, 6))
        spin_edit = Spinbox(frame_p, from_=1, to=999, width=6, font=FONT_UI)
        spin_edit.delete(0, tk.END)
        spin_edit.insert(0, str(vals[5]))
        spin_edit.pack(side="left")

        def _guardar_edit():
            try:
                nueva_cant = int(spin_edit.get())
            except ValueError:
                return
            if nueva_cant < 1:
                return
            if db_editar_envio(envio_id, nueva_cant):
                messagebox.showinfo("Actualizado", "Cantidad actualizada y stock ajustado.", parent=win)
                popup.destroy()
                filtrar()
            else:
                messagebox.showerror("Error", "No se pudo editar el envío.", parent=win)

        frame_btns = tk.Frame(popup, bg=BG_MAIN)
        frame_btns.pack(pady=10)
        btn_guardar = tk.Button(frame_btns, text="Guardar", command=_guardar_edit)
        _estilo_btn(btn_guardar, primario=True)
        btn_guardar.pack(side="left", padx=4)
        btn_cancel = tk.Button(frame_btns, text="Cancelar", command=popup.destroy)
        _estilo_btn(btn_cancel, primario=False)
        btn_cancel.pack(side="left", padx=4)

    # Habilitar/deshabilitar botones según selección
    def _on_select(_=None):
        sel = tree_env.selection()
        if sel:
            vals = tree_env.item(sel[0], "values")
            es_anulado = (vals[6] == "Anulado")
            btn_anular.config(state="disabled" if es_anulado else "normal")
            btn_editar.config(state="disabled" if es_anulado else "normal")
        else:
            btn_anular.config(state="disabled")
            btn_editar.config(state="disabled")

    tree_env.bind("<<TreeviewSelect>>", _on_select)

    btn_registrar.config(command=registrar)
    btn_filtrar.config(command=filtrar)
    btn_todos.config(command=mostrar_todos)
    btn_anular.config(command=anular_envio)
    btn_editar.config(command=editar_envio)

    mostrar_todos()

# ---------------------------------------------------------------------------
# Historial de monitoreos
# ---------------------------------------------------------------------------

def abrir_historial():
    """Abre la ventana de historial de monitoreos con filtros, ordenamiento, paginación y coloreo."""
    win = tk.Toplevel()
    win.title("Historial de Monitoreos")
    win.geometry("940x620")
    win.resizable(True, True)

    # ── Estado ──────────────────────────────────────────────────────────────
    all_filas    = []
    page_size    = 200
    current_page = 0
    sort_col     = None
    sort_asc     = True
    modo_arbol   = False
    grupos_por_pagina = 20

    # ── Fila 0: Filtros ────────────────────────────────────────────────────
    lbl_kw = {"bg": BG_MAIN, "font": FONT_UI, "fg": "#555555"}

    frame_filtros = tk.Frame(win, bg=BG_MAIN)
    frame_filtros.pack(fill="x", padx=10, pady=(10, 4))

    # Row 0: Desde / Hasta / Sucursal / Modelo
    row0 = tk.Frame(frame_filtros, bg=BG_MAIN)
    row0.pack(fill="x")
    tk.Label(row0, text="Desde:", **lbl_kw).pack(side="left", padx=(0, 4))
    var_desde = tk.StringVar(value=(datetime.today() - timedelta(days=30)).strftime("%d/%m/%Y"))
    tk.Entry(row0, textvariable=var_desde, width=11, font=FONT_UI).pack(side="left", padx=(0, 10))

    tk.Label(row0, text="Hasta:", **lbl_kw).pack(side="left", padx=(0, 4))
    var_hasta = tk.StringVar(value=datetime.today().strftime("%d/%m/%Y"))
    tk.Entry(row0, textvariable=var_hasta, width=11, font=FONT_UI).pack(side="left", padx=(0, 10))

    tk.Label(row0, text="Sucursal:", **lbl_kw).pack(side="left", padx=(0, 4))
    var_suc = tk.StringVar(value="Todas")
    suc_list = ["Todas"] + db_sucursales_activas()
    Combobox(row0, textvariable=var_suc, values=suc_list, state="readonly",
             width=14, font=FONT_UI).pack(side="left", padx=(0, 10))

    tk.Label(row0, text="Modelo:", **lbl_kw).pack(side="left", padx=(0, 4))
    var_modelo = tk.StringVar(value="Todos")
    mod_list = ["Todos"] + [n for (_, n) in db_modelos_listar()]
    Combobox(row0, textvariable=var_modelo, values=mod_list, state="readonly",
             width=18, font=FONT_UI).pack(side="left")

    # Row 1: IP / Estado Alerta / Botones
    row1 = tk.Frame(frame_filtros, bg=BG_MAIN)
    row1.pack(fill="x", pady=(4, 0))
    tk.Label(row1, text="IP:", **lbl_kw).pack(side="left", padx=(0, 4))
    var_ip = tk.StringVar(value="Todas")
    ip_list = ["Todas"] + [imp["ip"] for imp in db_impresoras_todas()]
    Combobox(row1, textvariable=var_ip, values=ip_list, state="readonly",
             width=15, font=FONT_UI).pack(side="left", padx=(0, 10))

    tk.Label(row1, text="Alerta:", **lbl_kw).pack(side="left", padx=(0, 4))
    var_alerta = tk.StringVar(value="Todos")
    Combobox(row1, textvariable=var_alerta, values=["Todos", "Bajo", "Medio", "Sin datos", "Normal"],
             state="readonly", width=10, font=FONT_UI).pack(side="left", padx=(0, 10))

    btn_filtrar_h = tk.Button(row1, text="Filtrar")
    btn_filtrar_h.pack(side="left", padx=(0, 4))
    btn_todos_h   = tk.Button(row1, text="Ver todos")
    btn_todos_h.pack(side="left", padx=(0, 4))
    btn_exportar_h = tk.Button(row1, text="Exportar Excel")
    btn_exportar_h.pack(side="left")
    for b in (btn_filtrar_h, btn_todos_h, btn_exportar_h):
        _estilo_btn(b, primario=False)

    # ── Fila 2: Tabla ─────────────────────────────────────────────────────
    cols_hist    = ("Fecha", "Sucursal", "IP", "Modelo", "Tóner (%)", "Unidad Imagen (%)", "Kit Mant. (%)")
    col_widths_h = (130, 110, 105, 145, 80, 110, 105)

    frame_tree = tk.Frame(win)
    frame_tree.pack(fill="both", expand=True, padx=10, pady=(4, 0))

    tree_h = Treeview(frame_tree, columns=cols_hist, show="headings", height=14)
    tree_h.column("#0", width=110, minwidth=70)
    for col, w in zip(cols_hist, col_widths_h):
        tree_h.heading(col, text=col)
        tree_h.column(col, anchor="center", width=w)
    tree_h.tag_configure("bajo",      background=COLOR_BAJO)
    tree_h.tag_configure("medio",     background=COLOR_MEDIO)
    tree_h.tag_configure("sin_datos", background=COLOR_SIN_DATOS)
    tree_h.tag_configure("fecha_grupo", font=("Segoe UI", 9, "bold"), foreground="#336699")
    sb_h = Scrollbar(frame_tree, orient="vertical", command=tree_h.yview)
    tree_h.configure(yscroll=sb_h.set)
    tree_h.pack(side="left", fill="both", expand=True)
    sb_h.pack(side="right", fill="y")

    # ── Fila 3: Barra inferior ────────────────────────────────────────────
    frame_pie = tk.Frame(win)
    frame_pie.pack(fill="x", padx=10, pady=(4, 8))

    btn_grafico_h = tk.Button(frame_pie, text="Ver Gráfico", state="disabled")
    btn_grafico_h.pack(side="left", padx=(0, 4))
    btn_catalogo_h = tk.Button(frame_pie, text="Ver en catálogo", state="disabled")
    btn_catalogo_h.pack(side="left", padx=(0, 10))

    btn_prev = tk.Button(frame_pie, text="<<", width=3)
    btn_prev.pack(side="left", padx=(0, 2))
    lbl_pag = tk.Label(frame_pie, text="", width=16, anchor="center", fg="gray")
    lbl_pag.pack(side="left")
    btn_next = tk.Button(frame_pie, text=">>", width=3)
    btn_next.pack(side="left", padx=(2, 0))

    lbl_total_h = tk.Label(frame_pie, text="", fg="gray")
    lbl_total_h.pack(side="right")

    btn_arbol = tk.Button(frame_pie, text="🌲 Vista árbol")
    btn_arbol.pack(side="right", padx=(6, 0))

    for b in (btn_grafico_h, btn_catalogo_h, btn_prev, btn_next, btn_arbol):
        _estilo_btn(b, primario=False)

    # ── Funciones ──────────────────────────────────────────────────────────
    def _fecha_a_db(s):
        s = s.strip()
        if not s:
            return None
        try:
            return datetime.strptime(s, "%d/%m/%Y").strftime("%Y-%m-%d")
        except ValueError:
            return None

    def mostrar_pagina():
        tree_h.delete(*tree_h.get_children())
        if modo_arbol:
            _mostrar_arbol()
        else:
            _mostrar_plano()

    def _mostrar_plano():
        inicio = current_page * page_size
        pagina = all_filas[inicio:inicio + page_size]
        for f in pagina:
            tree_h.insert("", "end", values=f[:7], tags=(f[7],))
        total = len(all_filas)
        pag_tot = max(1, (total + page_size - 1) // page_size)
        lbl_pag.config(text=f"Pág. {current_page + 1} / {pag_tot}")
        btn_prev.config(state="normal" if current_page > 0 else "disabled")
        btn_next.config(state="normal" if inicio + page_size < total else "disabled")
        lbl_total_h.config(text=f"Total: {total} registro{'s' if total != 1 else ''}")

    def _mostrar_arbol():
        nonlocal sort_col, sort_asc
        grupos = {}
        for f in all_filas:
            dia = f[0].split(" ")[0]
            grupos.setdefault(dia, []).append(f)
        try:
            fechas_ord = sorted(grupos.keys(),
                                key=lambda d: datetime.strptime(d, "%d/%m/%Y"),
                                reverse=not sort_asc if sort_col == 0 else False)
        except (ValueError, TypeError):
            fechas_ord = sorted(grupos.keys(), reverse=not sort_asc if sort_col == 0 else False)
        total_grupos = len(fechas_ord)
        inicio = current_page * grupos_por_pagina
        pagina = fechas_ord[inicio:inicio + grupos_por_pagina]
        for dia in pagina:
            regs = grupos[dia]
            n = len(regs)
            padre = tree_h.insert("", "end", text=f"  {dia}",
                                  values=("", "", "", f"{n} registro{'s' if n!=1 else ''}", "", "", ""),
                                  tags=("fecha_grupo",), open=False)
            for f in regs:
                tree_h.insert(padre, "end", text="", values=f[:7], tags=(f[7],))
        pag_tot = max(1, (total_grupos + grupos_por_pagina - 1) // grupos_por_pagina)
        lbl_pag.config(text=f"Pág. {current_page+1}/{pag_tot} ({total_grupos} fechas)")
        btn_prev.config(state="normal" if current_page > 0 else "disabled")
        btn_next.config(state="normal" if inicio + grupos_por_pagina < total_grupos else "disabled")
        lbl_total_h.config(text=f"Total: {len(all_filas)} regs, {total_grupos} fechas")

    def cargar():
        nonlocal all_filas, current_page, sort_col, sort_asc
        current_page = 0
        sort_col = None
        sort_asc = True
        for i, col in enumerate(cols_hist):
            tree_h.heading(col, text=col)

        filas = db_cargar_historial(
            desde    = _fecha_a_db(var_desde.get()),
            hasta    = _fecha_a_db(var_hasta.get()),
            sucursal = var_suc.get(),
            modelo   = var_modelo.get(),
            ip       = var_ip.get(),
        )
        alerta_filtro = var_alerta.get()
        if alerta_filtro != "Todos":
            tag_map = {"Bajo": "bajo", "Medio": "medio", "Sin datos": "sin_datos", "Normal": ""}
            target_tag = tag_map.get(alerta_filtro, "")
            filas = [f for f in filas if f[7] == target_tag]
        all_filas = filas
        mostrar_pagina()

    def filtrar_h():
        cargar()

    def todos_h():
        var_desde.set("")
        var_hasta.set("")
        var_suc.set("Todas")
        var_modelo.set("Todos")
        var_ip.set("Todas")
        var_alerta.set("Todos")
        cargar()

    def grafico_h():
        sel = tree_h.selection()
        if sel and not tree_h.get_children(sel[0]):
            mostrar_grafico(tree_h.item(sel[0], "values")[2])

    def ver_en_catalogo_h():
        sel = tree_h.selection()
        if sel and not tree_h.get_children(sel[0]):
            abrir_catalogo_impresoras(seleccionar_ip=tree_h.item(sel[0], "values")[2])

    def _sort_key(fila, col_idx):
        val = fila[col_idx]
        if col_idx == 0:
            try:
                return datetime.strptime(val, "%d/%m/%Y %H:%M")
            except (ValueError, TypeError):
                return datetime.min
        if col_idx in (4, 5, 6):
            try:
                return float(val.replace("%", "").replace(",", "."))
            except (ValueError, TypeError):
                return -1.0
        return val.lower()

    def ordenar_por_columna(col_idx):
        nonlocal sort_col, sort_asc
        if modo_arbol and col_idx != 0:
            return
        if sort_col == col_idx:
            sort_asc = not sort_asc
        else:
            sort_col = col_idx
            sort_asc = True
        if not modo_arbol:
            all_filas.sort(key=lambda f: _sort_key(f, col_idx), reverse=not sort_asc)
        for i, col in enumerate(cols_hist):
            arrow = " ▲" if sort_asc else " ▼"
            tree_h.heading(col, text=col + (arrow if i == col_idx else ""))
        mostrar_pagina()

    def _total_items():
        if modo_arbol:
            if not all_filas:
                return 0
            dias = {f[0].split(" ")[0] for f in all_filas}
            return len(dias)
        return len(all_filas)

    def pagina_siguiente():
        nonlocal current_page
        paso = grupos_por_pagina if modo_arbol else page_size
        items = _total_items()
        if (current_page + 1) * paso < items:
            current_page += 1
            mostrar_pagina()

    def pagina_anterior():
        nonlocal current_page
        if current_page > 0:
            current_page -= 1
            mostrar_pagina()

    def exportar_hist():
        if not all_filas:
            messagebox.showwarning("Sin datos", "No hay filas para exportar.", parent=win)
            return
        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"historial_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            parent=win)
        if not ruta:
            return
        wb = Workbook(); ws = wb.active; ws.title = "Historial"
        headers = list(cols_hist)
        widths  = [17, 18, 14, 24, 11, 16, 14]
        hdr_fill = PatternFill("solid", fgColor="CCCCCC")
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hdr_fill; c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[c.column_letter].width = w
        fills = {"bajo":     PatternFill("solid", fgColor="FF9999"),
                 "medio":    PatternFill("solid", fgColor="FFFF99"),
                 "sin_datos":PatternFill("solid", fgColor="D0D0D0")}
        for f in all_filas:
            ws.append(list(f[:7]))
            tag = f[7]
            if tag in fills:
                for col in range(1, 8):
                    ws.cell(row=ws.max_row, column=col).fill = fills[tag]
        wb.save(ruta)
        messagebox.showinfo("Exportado", f"Historial guardado:\n{ruta}", parent=win)

    # ── Comandos ───────────────────────────────────────────────────────────
    btn_filtrar_h.config(command=filtrar_h)
    btn_todos_h.config(command=todos_h)
    btn_grafico_h.config(command=grafico_h)
    btn_catalogo_h.config(command=ver_en_catalogo_h)
    btn_exportar_h.config(command=exportar_hist)
    btn_prev.config(command=pagina_anterior)
    btn_next.config(command=pagina_siguiente)

    # ── Filtro automático ──────────────────────────────────────────────────
    var_desde.trace_add("write",  lambda *_: win.after(400, cargar))
    var_hasta.trace_add("write",  lambda *_: win.after(400, cargar))
    var_suc.trace_add("write",    lambda *_: cargar())
    var_modelo.trace_add("write", lambda *_: cargar())
    var_ip.trace_add("write",     lambda *_: cargar())
    var_alerta.trace_add("write", lambda *_: cargar())

    # ── Ordenamiento por columna ───────────────────────────────────────────
    for i, col in enumerate(cols_hist):
        tree_h.heading(col, text=col, command=lambda i=i: ordenar_por_columna(i))

    # ── Doble clic → gráfico ──────────────────────────────────────────────
    tree_h.bind("<<TreeviewDoubleClick>>", lambda _: grafico_h())

    # ── Selección → habilitar botones ─────────────────────────────────────
    def _on_select_h(_=None):
        sel = bool(tree_h.selection())
        btn_grafico_h.config(state="normal" if sel else "disabled")
        btn_catalogo_h.config(state="normal" if sel else "disabled")
    tree_h.bind("<<TreeviewSelect>>", _on_select_h)

    def toggle_arbol():
        nonlocal modo_arbol, current_page
        modo_arbol = not modo_arbol
        current_page = 0
        btn_arbol.config(text="📋 Vista plana" if modo_arbol else "🌲 Vista árbol")
        tree_h.config(show="tree headings" if modo_arbol else "headings")
        # Resetear heading names quitando flechas de ordenamiento
        for i, col in enumerate(cols_hist):
            tree_h.heading(col, text=col)
        mostrar_pagina()

    btn_arbol.config(command=toggle_arbol)

    # ── Carga inicial ──────────────────────────────────────────────────────
    cargar()

# ---------------------------------------------------------------------------
# Catálogo de impresoras (CRUD)
# ---------------------------------------------------------------------------

def _dialogo_impresora(parent, titulo, valores_iniciales=None):
    """Muestra un diálogo para agregar/editar una impresora.
    valores_iniciales: dict con claves ip, modelo, sucursal, nombre, sn, ubicacion.
    Retorna dict o None si se canceló.
    """
    dlg = tk.Toplevel(parent)
    dlg.title(titulo)
    dlg.geometry("380x310")
    dlg.resizable(False, False)
    dlg.grab_set()

    modelos_lista = [n for (_, n) in db_modelos_listar()]
    vals = valores_iniciales or {}

    fields = [
        ("IP:",       "ip",       dict(width=22)),
        ("Sucursal:", "sucursal", dict(width=22)),
        ("Ubicación:","ubicacion",dict(width=22)),
        ("Nombre:",   "nombre",   dict(width=22)),
        ("N° Serie:", "sn",       dict(width=22)),
    ]
    vars_ = {}
    for r, (lbl, key, opts) in enumerate(fields):
        tk.Label(dlg, text=lbl, anchor="e", width=10).grid(row=r, column=0, padx=10, pady=4, sticky="e")
        var = tk.StringVar(value=vals.get(key, ""))
        tk.Entry(dlg, textvariable=var, **opts).grid(row=r, column=1, padx=(0, 10), sticky="w")
        vars_[key] = var

    tk.Label(dlg, text="Modelo:", anchor="e", width=10).grid(row=5, column=0, padx=10, pady=4, sticky="e")
    var_modelo = tk.StringVar(value=vals.get("modelo", modelos_lista[0] if modelos_lista else ""))
    Combobox(dlg, textvariable=var_modelo, values=modelos_lista,
             state="readonly", width=20).grid(row=5, column=1, padx=(0, 10), sticky="w")
    vars_["modelo"] = var_modelo

    resultado = [None]

    def aceptar():
        ip       = vars_["ip"].get().strip()
        modelo   = vars_["modelo"].get().strip()
        sucursal = vars_["sucursal"].get().strip()
        if not ip:
            messagebox.showwarning("Campo requerido", "La IP es obligatoria.", parent=dlg)
            return
        if not modelo:
            messagebox.showwarning("Campo requerido", "Seleccione un modelo.", parent=dlg)
            return
        resultado[0] = {
            "ip":       ip,
            "modelo":   modelo,
            "sucursal": sucursal,
            "ubicacion":vars_["ubicacion"].get().strip(),
            "nombre":   vars_["nombre"].get().strip(),
            "sn":       vars_["sn"].get().strip(),
        }
        dlg.destroy()

    frame_btns = tk.Frame(dlg)
    frame_btns.grid(row=6, column=0, columnspan=2, pady=10)
    tk.Button(frame_btns, text="Aceptar",  width=10, command=aceptar).pack(side="left", padx=6)
    tk.Button(frame_btns, text="Cancelar", width=10, command=dlg.destroy).pack(side="left", padx=6)

    parent.wait_window(dlg)
    return resultado[0]


def abrir_gestion_modelos(parent=None):
    """Ventana para gestionar (agregar, renombrar, eliminar) modelos de impresoras."""
    win = tk.Toplevel(parent)
    win.title("Gestionar Modelos")
    win.geometry("420x320")
    win.resizable(False, False)
    if parent:
        win.transient(parent)
    win.grab_set()

    def refrescar():
        lb.delete(0, "end")
        for mid, mnombre in db_modelos_listar():
            lb.insert("end", mnombre)

    frame_top = tk.Frame(win)
    frame_top.pack(fill="x", padx=10, pady=(10, 4))

    tk.Label(frame_top, text="Modelos existentes:").pack(anchor="w")
    lb = tk.Listbox(frame_top, height=10)
    lb.pack(fill="both", expand=True)

    frame_btns = tk.Frame(win)
    frame_btns.pack(fill="x", padx=10, pady=(8, 10))

    def agregar():
        nombre = simpledialog.askstring("Agregar Modelo", "Nombre del modelo:", parent=win)
        if not nombre:
            return
        nombre = nombre.strip()
        if not nombre:
            return
        try:
            db_modelo_agregar(nombre)
            refrescar()
        except sqlite3.IntegrityError:
            messagebox.showwarning("Duplicado", f"El modelo '{nombre}' ya existe.", parent=win)

    def renombrar():
        sel = lb.curselection()
        if not sel:
            messagebox.showinfo("Sin selección", "Seleccione un modelo.", parent=win)
            return
        nombre_actual = lb.get(sel[0])
        nuevo = simpledialog.askstring("Renombrar Modelo",
                                       f"Nuevo nombre para '{nombre_actual}':",
                                       parent=win, initialvalue=nombre_actual)
        if not nuevo:
            return
        nuevo = nuevo.strip()
        if not nuevo or nuevo == nombre_actual:
            return
        modelos = db_modelos_listar()
        mid = modelos[sel[0]][0]
        try:
            db_modelo_renombrar(mid, nuevo)
            refrescar()
        except sqlite3.IntegrityError:
            messagebox.showwarning("Duplicado", f"Ya existe un modelo '{nuevo}'.", parent=win)

    def eliminar():
        sel = lb.curselection()
        if not sel:
            messagebox.showinfo("Sin selección", "Seleccione un modelo.", parent=win)
            return
        nombre = lb.get(sel[0])
        if not messagebox.askyesno("Confirmar",
                                   f"¿Eliminar modelo '{nombre}'?\nSolo se eliminará si ninguna impresora lo usa.",
                                   parent=win):
            return
        modelos = db_modelos_listar()
        mid = modelos[sel[0]][0]
        try:
            db_modelo_eliminar(mid)
            refrescar()
        except ValueError as e:
            messagebox.showwarning("En uso", str(e), parent=win)

    btn_agregar   = tk.Button(frame_btns, text="Agregar",   command=agregar)
    btn_renombrar = tk.Button(frame_btns, text="Renombrar", command=renombrar)
    btn_eliminar  = tk.Button(frame_btns, text="Eliminar",  command=eliminar)
    btn_cerrar    = tk.Button(frame_btns, text="Cerrar",    command=win.destroy)
    for b in (btn_agregar, btn_renombrar, btn_eliminar, btn_cerrar):
        b.pack(side="left", padx=4)
        _estilo_btn(b, primario=False)

    refrescar()
    parent.wait_window(win)


def _importar_desde_excel(parent):
    """Importa impresoras desde un Excel (columnas D=IP, E=Modelo)."""
    archivo = filedialog.askopenfilename(
        parent=parent,
        filetypes=[("Archivos Excel", "*.xlsx")],
        title="Seleccionar Excel de impresoras",
    )
    if not archivo:
        return

    try:
        wb = load_workbook(archivo, data_only=True)
        ws = wb.active
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{e}", parent=parent)
        return

    importadas = 0
    omitidas   = 0
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        if len(row) < 5:
            continue
        ip     = str(row[3]).strip() if row[3] else ""
        modelo = str(row[4]).strip() if row[4] else ""
        if not ip or not modelo:
            continue
        try:
            db_impresora_agregar(ip, modelo, sucursal="")
            importadas += 1
        except sqlite3.IntegrityError:
            omitidas += 1

    messagebox.showinfo(
        "Importación completada",
        f"Impresoras importadas: {importadas}\nOmitidas (IP duplicada): {omitidas}",
        parent=parent,
    )


def abrir_catalogo_impresoras(seleccionar_ip=None):
    """Abre la ventana de gestión del catálogo de impresoras."""
    win = tk.Toplevel()
    win.title("Catálogo de Impresoras")
    win.geometry("920x620")
    win.resizable(True, True)

    # Mapeo iid → id de DB
    id_map = {}
    # Mapeo ip → iid
    ip_to_iid = {}
    # Estado de ordenamiento
    sort_col = None
    sort_asc = True
    # Cache de último tóner
    toner_cache = {}

    def _cargar_toner_cache():
        toner_cache.clear()
        todas = db_impresoras_todas()
        for imp in todas:
            nivel = db_ultimo_toner(imp["ip"])
            if nivel is not None:
                toner_cache[imp["ip"]] = f"{nivel*100:.0f}%"

    _cargar_toner_cache()

    # ------------------------------------------------------------------
    # Fila 0: Filtros
    # ------------------------------------------------------------------
    frame_top = tk.Frame(win)
    frame_top.pack(fill="x", padx=10, pady=(10, 4))

    tk.Label(frame_top, text="Buscar:").pack(side="left", padx=(0, 4))
    var_buscar = tk.StringVar()
    tk.Entry(frame_top, textvariable=var_buscar, width=14).pack(side="left", padx=(0, 6))

    tk.Label(frame_top, text="Modelo:").pack(side="left", padx=(0, 4))
    var_filtro_modelo = tk.StringVar(value="Todos")
    modelos_filtro = ["Todos"] + db_modelos_activos()
    Combobox(frame_top, textvariable=var_filtro_modelo, values=modelos_filtro,
             state="readonly", width=15).pack(side="left", padx=(0, 6))

    tk.Label(frame_top, text="Sucursal:").pack(side="left", padx=(0, 4))
    var_filtro_suc = tk.StringVar(value="Todas")
    suc_filtro = ["Todas"] + db_sucursales_activas()
    Combobox(frame_top, textvariable=var_filtro_suc, values=suc_filtro,
             state="readonly", width=12).pack(side="left", padx=(0, 6))

    tk.Label(frame_top, text="Estado:").pack(side="left", padx=(0, 4))
    var_estado = tk.StringVar(value="Activas")
    Combobox(frame_top, textvariable=var_estado, values=["Todas", "Activas", "Inactivas"],
             state="readonly", width=8).pack(side="left", padx=(0, 6))

    btn_filtrar_cat = tk.Button(frame_top, text="Filtrar")
    btn_filtrar_cat.pack(side="left", padx=(0, 4))
    btn_todas_cat = tk.Button(frame_top, text="Todas")
    btn_todas_cat.pack(side="left")

    # ------------------------------------------------------------------
    # Fila 1: Tabla
    # ------------------------------------------------------------------
    cols_cat   = ("Estado", "Sucursal", "Ubicación", "IP", "Modelo", "Nombre", "N° Serie", "Tóner")
    col_widths = (55, 110, 90, 110, 130, 95, 85, 55)

    frame_tree = tk.Frame(win)
    frame_tree.pack(fill="both", expand=True, padx=10, pady=(4, 0))

    tree_cat = Treeview(frame_tree, columns=cols_cat, show="headings",
                        height=12, selectmode="extended")
    for col, w in zip(cols_cat, col_widths):
        tree_cat.heading(col, text=col)
        tree_cat.column(col, anchor="center" if col in ("Estado", "IP", "N° Serie", "Tóner") else "w", width=w)
    tree_cat.tag_configure("baja", foreground="#999999")
    sb_cat = Scrollbar(frame_tree, orient="vertical", command=tree_cat.yview)
    tree_cat.configure(yscroll=sb_cat.set)
    tree_cat.pack(side="left", fill="both", expand=True)
    sb_cat.pack(side="right", fill="y")

    # ------------------------------------------------------------------
    # Fila 2: Panel resumen
    # ------------------------------------------------------------------
    frame_resumen = tk.Frame(win)
    frame_resumen.pack(fill="x", padx=10, pady=(4, 2))
    lbl_total = tk.Label(frame_resumen, text="", font=FONT_UI, fg="#555555")
    lbl_total.pack(side="left")

    # ------------------------------------------------------------------
    # Fila 3: Botones de acción
    # ------------------------------------------------------------------
    frame_btns = tk.Frame(win)
    frame_btns.pack(fill="x", padx=10, pady=(4, 10))

    btn_agregar = tk.Button(frame_btns, text="Agregar")
    btn_agregar.pack(side="left", padx=4)
    btn_editar = tk.Button(frame_btns, text="Editar")
    btn_editar.pack(side="left", padx=4)
    btn_duplicar = tk.Button(frame_btns, text="Duplicar")
    btn_duplicar.pack(side="left", padx=4)
    btn_baja = tk.Button(frame_btns, text="Dar de baja / Reactivar")
    btn_baja.pack(side="left", padx=4)
    btn_importar = tk.Button(frame_btns, text="Importar Excel")
    btn_importar.pack(side="left", padx=4)
    btn_exportar = tk.Button(frame_btns, text="Exportar Excel")
    btn_exportar.pack(side="left", padx=4)
    btn_modelos = tk.Button(frame_btns, text="Gestionar Modelos")
    btn_modelos.pack(side="left", padx=4)

    # Aplicar estilo a botones
    for b in (btn_agregar, btn_editar, btn_duplicar, btn_baja, btn_importar, btn_exportar,
              btn_modelos, btn_filtrar_cat, btn_todas_cat):
        _estilo_btn(b, primario=False)

    # ------------------------------------------------------------------
    # Funciones internas
    # ------------------------------------------------------------------
    def cargar_catalogo(solo_activas=None, texto="", filtro_modelo="", filtro_sucursal=""):
        tree_cat.delete(*tree_cat.get_children())
        id_map.clear()
        ip_to_iid.clear()
        t = texto.lower()
        activas_cnt = inactivas_cnt = 0
        for imp in db_impresoras_todas():
            if solo_activas is True and not imp["activa"]:
                continue
            if solo_activas is False and imp["activa"]:
                continue
            if filtro_modelo and filtro_modelo != "Todos" and imp["modelo"] != filtro_modelo:
                continue
            if filtro_sucursal and filtro_sucursal != "Todas" and imp["sucursal"] != filtro_sucursal:
                continue
            if t and t not in str(imp["ip"]).lower() and t not in str(imp["sucursal"]).lower() \
                 and t not in str(imp["modelo"]).lower() and t not in str(imp["nombre"] or "").lower() \
                 and t not in str(imp.get("sn", "") or "").lower() \
                 and t not in str(imp.get("ubicacion", "") or "").lower():
                continue
            estado_str = "Activa" if imp["activa"] else "Baja"
            toner_str = toner_cache.get(imp["ip"], "—")
            tags = ("baja",) if not imp["activa"] else ()
            iid = tree_cat.insert("", "end", values=(
                estado_str, imp["sucursal"], imp.get("ubicacion", "") or "",
                imp["ip"], imp["modelo"], imp["nombre"] or "",
                imp.get("sn", "") or "", toner_str
            ), tags=tags)
            id_map[iid] = imp["id"]
            ip_to_iid[imp["ip"]] = iid
            if imp["activa"]:
                activas_cnt += 1
            else:
                inactivas_cnt += 1
        lbl_total.config(
            text=f"Total: {activas_cnt + inactivas_cnt}  |  Activas: {activas_cnt}  |  Inactivas: {inactivas_cnt}")

    def filtrar_cat():
        estado = var_estado.get()
        solo = None if estado == "Todas" else (estado == "Activas")
        cargar_catalogo(
            solo_activas=solo, texto=var_buscar.get().strip(),
            filtro_modelo=var_filtro_modelo.get(),
            filtro_sucursal=var_filtro_suc.get())

    def mostrar_todas_cat():
        var_buscar.set("")
        var_estado.set("Todas")
        var_filtro_modelo.set("Todos")
        var_filtro_suc.set("Todas")
        cargar_catalogo()

    def ordenar_por_columna(col_idx):
        nonlocal sort_col, sort_asc
        if sort_col == col_idx:
            sort_asc = not sort_asc
        else:
            sort_col = col_idx
            sort_asc = True
        items = [(tree_cat.set(iid, col_idx), iid) for iid in tree_cat.get_children("")]
        try:
            items.sort(key=lambda x: (x[0] == "—", float(x[0].replace("%", "")) if x[0].replace("%", "").replace(".", "").isdigit() else x[0].lower()),
                       reverse=not sort_asc)
        except (ValueError, TypeError):
            items.sort(key=lambda x: x[0].lower(), reverse=not sort_asc)
        for idx, (_, iid) in enumerate(items):
            tree_cat.move(iid, "", idx)
        for i, col in enumerate(cols_cat):
            arrow = " ▲" if sort_asc else " ▼"
            tree_cat.heading(col, text=col + (arrow if i == col_idx else ""))

    def agregar():
        vals = _dialogo_impresora(win, "Agregar Impresora")
        if vals:
            try:
                db_impresora_agregar(vals["ip"], vals["modelo"], vals["sucursal"],
                                     vals["nombre"], vals["sn"], vals["ubicacion"])
                _cargar_toner_cache()
                filtrar_cat()
            except sqlite3.IntegrityError:
                messagebox.showwarning(
                    "IP duplicada", f"Ya existe una impresora con la IP {vals['ip']}.", parent=win)

    def editar():
        sel = tree_cat.selection()
        if not sel:
            messagebox.showinfo("Sin selección", "Seleccione una impresora para editar.", parent=win)
            return
        iid = sel[0]
        id_ = id_map[iid]
        vals_actuales = tree_cat.item(iid, "values")
        vals = _dialogo_impresora(win, "Editar Impresora", {
            "ip":       vals_actuales[3],
            "modelo":   vals_actuales[4],
            "sucursal": vals_actuales[1],
            "ubicacion":vals_actuales[2],
            "nombre":   vals_actuales[5],
            "sn":       vals_actuales[6],
        })
        if vals:
            activa = (vals_actuales[0] == "Activa")
            try:
                db_impresora_actualizar(id_, vals["ip"], vals["modelo"],
                                        vals["sucursal"], vals["nombre"], vals["sn"], activa,
                                        vals["ubicacion"])
                _cargar_toner_cache()
                filtrar_cat()
            except sqlite3.IntegrityError:
                messagebox.showwarning(
                    "IP duplicada", f"Ya existe una impresora con la IP {vals['ip']}.", parent=win)

    def duplicar():
        sel = tree_cat.selection()
        if not sel:
            messagebox.showinfo("Sin selección", "Seleccione una impresora para duplicar.", parent=win)
            return
        vals_actuales = tree_cat.item(sel[0], "values")
        vals = _dialogo_impresora(win, "Duplicar Impresora", {
            "ip":       "",
            "modelo":   vals_actuales[4],
            "sucursal": vals_actuales[1],
            "ubicacion":vals_actuales[2],
            "nombre":   vals_actuales[5],
            "sn":       "",
        })
        if vals:
            if not vals["ip"]:
                messagebox.showwarning("IP requerida", "Debe ingresar una IP diferente.", parent=win)
                return
            try:
                db_impresora_agregar(vals["ip"], vals["modelo"], vals["sucursal"],
                                     vals["nombre"], vals["sn"], vals["ubicacion"])
                filtrar_cat()
            except sqlite3.IntegrityError:
                messagebox.showwarning(
                    "IP duplicada", f"Ya existe una impresora con la IP {vals['ip']}.", parent=win)

    def alternar_baja():
        sel = tree_cat.selection()
        if not sel:
            messagebox.showinfo("Sin selección", "Seleccione una o más impresoras.", parent=win)
            return
        # Ver si todas son del mismo tipo (todas activas o todas inactivas)
        estados = {tree_cat.item(iid, "values")[0] for iid in sel}
        if len(estados) > 1:
            messagebox.showwarning("Estado mixto",
                                   "Seleccione solo impresoras activas o solo inactivas.", parent=win)
            return
        es_activa = ("Activa" in estados)
        accion = "dar de baja" if es_activa else "reactivar"
        if not messagebox.askyesno("Confirmar",
                                   f"¿{accion.capitalize()} {len(sel)} impresora(s)?", parent=win):
            return
        todos = db_impresoras_todas()
        for iid in sel:
            id_ = id_map[iid]
            imp = next((f for f in todos if f["id"] == id_), None)
            if imp:
                db_impresora_actualizar(id_, imp["ip"], imp["modelo"], imp["sucursal"],
                                        imp["nombre"], imp.get("sn", ""), not es_activa,
                                        imp.get("ubicacion", ""))
        filtrar_cat()

    def importar_excel():
        _importar_desde_excel(win)
        _cargar_toner_cache()
        filtrar_cat()

    def exportar_cat():
        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"catalogo_impresoras_{datetime.now().strftime('%Y%m%d')}.xlsx",
            parent=win)
        if not ruta:
            return
        wb = Workbook(); ws = wb.active; ws.title = "Catálogo"
        headers = list(cols_cat)
        widths  = [10, 18, 14, 15, 24, 20, 14, 8]
        hdr_fill = PatternFill("solid", fgColor="4472C4")
        hdr_font = Font(bold=True, color="FFFFFF")
        align_center = Alignment(horizontal="center")
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = align_center
            ws.column_dimensions[c.column_letter].width = w
        baja_fill = PatternFill("solid", fgColor="D9D9D9")
        for imp in db_impresoras_todas():
            estado = "Activa" if imp["activa"] else "Baja"
            toner_str = toner_cache.get(imp["ip"], "—")
            ws.append([estado, imp["sucursal"], imp.get("ubicacion", "") or "",
                       imp["ip"], imp["modelo"], imp["nombre"] or "",
                       imp.get("sn", "") or "", toner_str])
            if not imp["activa"]:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col).fill = baja_fill
        wb.save(ruta)
        messagebox.showinfo("Exportado", f"Catálogo guardado:\n{ruta}", parent=win)

    # ── Asignar comandos ─────────────────────────────────────────────
    btn_filtrar_cat.config(command=filtrar_cat)
    btn_todas_cat.config(command=mostrar_todas_cat)
    btn_agregar.config(command=agregar)
    btn_editar.config(command=editar)
    btn_duplicar.config(command=duplicar)
    btn_baja.config(command=alternar_baja)
    btn_importar.config(command=importar_excel)
    btn_exportar.config(command=exportar_cat)
    btn_modelos.config(command=lambda: abrir_gestion_modelos(win))

    # ── Filtro automático en tiempo real ──
    var_buscar.trace_add("write", lambda *_: filtrar_cat())
    var_filtro_modelo.trace_add("write", lambda *_: filtrar_cat())
    var_filtro_suc.trace_add("write", lambda *_: filtrar_cat())
    var_estado.trace_add("write", lambda *_: filtrar_cat())

    # ── Ordenamiento por columna ──
    for i, col in enumerate(cols_cat):
        tree_cat.heading(col, text=col, command=lambda i=i: ordenar_por_columna(i))

    # ── Carga inicial (solo activas) ──
    cargar_catalogo(solo_activas=True)
    if seleccionar_ip and seleccionar_ip in ip_to_iid:
        target = ip_to_iid[seleccionar_ip]
        tree_cat.selection_set(target)
        tree_cat.see(target)
        tree_cat.focus(target)
        # Cambiar filtro a "Todas" para asegurar que la IP seleccionada sea visible
        var_estado.set("Todas")
        var_buscar.set(seleccionar_ip)
        cargar_catalogo(texto=seleccionar_ip)
        if seleccionar_ip in ip_to_iid:
            target = ip_to_iid[seleccionar_ip]
            tree_cat.selection_set(target)
            tree_cat.see(target)
            tree_cat.focus(target)

# ---------------------------------------------------------------------------
# Helpers de UI — filtro y ordenamiento
# ---------------------------------------------------------------------------

def aplicar_filtro(ctx):
    """Re-renderiza el Treeview aplicando el filtro de búsqueda y solo-alertas actuales."""
    texto        = ctx.entrada_busqueda.get().strip().lower()
    solo_alertas = ctx.var_solo_alertas.get()

    ctx.tree.delete(*ctx.tree.get_children())
    visible_idx = 0
    for fila in ctx.filas_tabla:
        sucursal, ip, modelo = fila[0], fila[1], fila[2]
        tag                  = fila[6]
        if texto and texto not in sucursal.lower() \
                 and texto not in ip.lower() \
                 and texto not in modelo.lower():
            continue
        if solo_alertas and tag not in ("bajo", "medio", "sin_datos"):
            continue
        parity = "par" if visible_idx % 2 == 0 else "impar"
        tags   = (tag,) if tag else (parity,)
        ctx.tree.insert("", "end", values=fila[:6], tags=tags)
        visible_idx += 1


def ordenar_por_columna(ctx, col_idx):
    """Ordena ctx.filas_tabla por la columna dada y re-aplica el filtro."""
    if ctx.sort_col == col_idx:
        ctx.sort_asc = not ctx.sort_asc
    else:
        ctx.sort_col = col_idx
        ctx.sort_asc = True

    def clave(fila):
        val = fila[col_idx]
        if isinstance(val, str) and val.endswith('%'):
            try:
                return float(val[:-1])
            except ValueError:
                return -1.0
        if val == "-":
            return -1.0
        if col_idx == 3:
            try:
                return datetime.strptime(val, "%d/%m/%Y %H:%M")
            except (ValueError, TypeError):
                return datetime.min
        return val.lower() if isinstance(val, str) else (val or "")

    ctx.filas_tabla.sort(key=clave, reverse=not ctx.sort_asc)
    aplicar_filtro(ctx)

    for i, col in enumerate(COLUMNAS_TREE):
        arrow = (" ▲" if ctx.sort_asc else " ▼") if i == ctx.sort_col else ""
        ctx.tree.heading(col, text=col + arrow)

# ---------------------------------------------------------------------------
# Monitoreo automático
# ---------------------------------------------------------------------------

def _iniciar_auto(ctx):
    """Arranca el countdown con el intervalo seleccionado en el combo."""
    seg = INTERVALOS_AUTO.get(ctx.combo_intervalo.get(), 3600)
    ctx.after_id = ctx.ventana.after(1000, _tick_auto, ctx, seg)


def _cancelar_auto(ctx):
    """Cancela el countdown pendiente y resetea el label."""
    if ctx.after_id is not None:
        ctx.ventana.after_cancel(ctx.after_id)
        ctx.after_id = None
    ctx.lbl_proximo.config(text="Próximo: --:--")


def _tick_auto(ctx, restantes):
    """Tick de 1 segundo. Al llegar a 0 lanza el monitoreo automático."""
    if not ctx.var_auto.get():
        return
    if restantes <= 0:
        if ctx.btn_iniciar["state"] == "normal":
            iniciar_monitoreo(ctx, es_automatico=True)
        else:
            # Monitoreo en curso: reintentar en 5 s
            ctx.after_id = ctx.ventana.after(5000, _tick_auto, ctx, 0)
        return
    mins, segs = divmod(restantes, 60)
    ctx.lbl_proximo.config(text=f"Próximo: {mins:02d}:{segs:02d}")
    ctx.after_id = ctx.ventana.after(1000, _tick_auto, ctx, restantes - 1)


# ---------------------------------------------------------------------------
# Helpers de UI  (siempre se ejecutan en el hilo principal vía ventana.after)
# ---------------------------------------------------------------------------

def _actualizar_progreso(ctx, sucursal, ip, modelo, toner, kit, unidad, progreso, umbral_bajo, umbral_medio):
    """Actualiza barra de progreso, etiqueta de estado e inserta la fila respetando el filtro."""
    ctx.resultado_label.config(text=f"Monitoreando: {ip} ({progreso}%)")
    ctx.barra_progreso.config(value=progreso)

    valores = [v for v in (toner, kit, unidad) if v is not None]
    tag     = clasificar_nivel(valores, umbral_bajo, umbral_medio)

    toner_str  = f"{toner*100:.1f}%"  if toner  is not None else "—"
    unidad_str = f"{unidad*100:.1f}%" if unidad is not None else "—"
    kit_str    = f"{kit*100:.1f}%"    if kit    is not None else "—"

    fila = (sucursal, ip, modelo, fecha_ult, toner_str, unidad_str, kit_str, tag)
    ctx.filas_tabla.append(fila)

    texto        = ctx.entrada_busqueda.get().strip().lower()
    solo_alertas = ctx.var_solo_alertas.get()
    if (not texto or texto in sucursal.lower() or texto in ip.lower() or texto in modelo.lower()) and \
       (not solo_alertas or tag in ("bajo", "medio", "sin_datos")):
        parity = "par" if (len(ctx.filas_tabla) - 1) % 2 == 0 else "impar"
        tags   = (tag,) if tag else (parity,)
        ctx.tree.insert("", "end", values=fila[:6], tags=tags)


def _finalizar(ctx, mensaje, tipo):
    """Resetea controles y muestra el mensaje final al terminar el monitoreo."""
    ctx.btn_iniciar.config(state="normal")
    ctx.btn_cancelar.config(state="disabled")
    ctx.btn_exportar.config(state="normal" if ctx.filas_tabla else "disabled")
    ctx.barra_progreso["value"] = 100 if tipo in ("exito", "alerta") else 0
    colores = {"exito": "green", "alerta": "orange", "cancelado": "orange", "error": "red"}
    ctx.resultado_label.config(text=mensaje, fg=colores.get(tipo, "red"))
    if tipo == "exito":
        messagebox.showinfo("Éxito", mensaje, parent=ctx.ventana)
    elif tipo == "alerta":
        messagebox.showwarning("Alertas de consumibles", mensaje, parent=ctx.ventana)
    elif tipo == "error":
        messagebox.showerror("Error", mensaje, parent=ctx.ventana)
    # Actualizar dashboard de stock
    _actualizar_stock_resumen(ctx)
    # Reiniciar countdown si el monitoreo automático está activo
    if tipo in ("exito", "alerta") and hasattr(ctx, "var_auto") and ctx.var_auto.get():
        _iniciar_auto(ctx)


def _actualizar_resumen(ctx, total, respondieron, sin_datos, bajo, medio, fecha_ult=""):
    """Actualiza el panel de resumen con las estadísticas del último monitoreo."""
    lr = ctx.labels_resumen
    lr["total"].config(text=str(total))
    lr["respondieron"].config(text=str(respondieron))
    lr["sin_datos"].config(text=str(sin_datos), fg="gray" if sin_datos == 0 else "black")
    lr["bajo"].config(text=str(bajo),   fg="red"    if bajo  > 0 else "black")
    lr["medio"].config(text=str(medio), fg="orange" if medio > 0 else "black")


def _actualizar_stock_resumen(ctx):
    """Actualiza los labels de stock crítico/bajo en el panel Resumen."""
    stock    = db_stock_obtener()
    criticos = sum(1 for r in stock if r["cantidad"] <= r["stock_minimo"])
    bajos    = sum(1 for r in stock if r["stock_minimo"] < r["cantidad"] <= r["stock_minimo"] * 2)
    ctx.labels_resumen["stock_critico"].config(
        text=str(criticos),
        fg=COLOR_BAJO if criticos > 0 else "#4CAF50")
    ctx.labels_resumen["stock_bajo"].config(
        text=str(bajos),
        fg=COLOR_MEDIO if bajos > 0 else "#4CAF50")

# ---------------------------------------------------------------------------
# Notificaciones por email
# ---------------------------------------------------------------------------

def enviar_alerta_email(impresoras_bajas, umbral_bajo):
    """Envía un email de alerta con las impresoras en nivel bajo. Corre en hilo secundario."""
    cfg          = cargar_config()
    remitente    = cfg.get("email_remitente", "").strip()
    password     = cfg.get("email_password", "")
    destinatarios_raw = cfg.get("email_destinatarios", "").strip()
    servidor     = cfg.get("email_servidor", "smtp.office365.com")
    puerto       = int(cfg.get("email_puerto", 587))

    if not remitente or not destinatarios_raw:
        _log.error("Email: remitente o destinatarios no configurados.")
        return

    destinatarios = [d.strip() for d in destinatarios_raw.split(",") if d.strip()]

    # Construir tabla HTML
    filas_html = ""
    for ip, modelo, sucursal, toner, kit, unidad in impresoras_bajas:
        t_str = f"{toner*100:.1f}%" if toner  is not None else "—"
        k_str = f"{kit*100:.1f}%"   if kit    is not None else "—"
        u_str = f"{unidad*100:.1f}%" if unidad is not None else "—"
        filas_html += (
            f"<tr><td>{sucursal}</td><td>{ip}</td><td>{modelo}</td>"
            f"<td style='color:red'>{t_str}</td>"
            f"<td style='color:red'>{u_str}</td>"
            f"<td style='color:red'>{k_str}</td></tr>"
        )

    html = f"""
    <html><body>
    <h2 style="color:#cc0000;">⚠ Alerta de consumibles — Nivel bajo (&lt;{umbral_bajo}%)</h2>
    <p>Las siguientes impresoras tienen uno o más consumibles por debajo del umbral configurado:</p>
    <table border="1" cellpadding="5" cellspacing="0" style="border-collapse:collapse;font-family:Arial,sans-serif;font-size:13px">
      <thead style="background:#f2f2f2">
        <tr><th>Sucursal</th><th>IP</th><th>Modelo</th>
            <th>Tóner</th><th>Unidad Imagen</th><th>Kit Mant.</th></tr>
      </thead>
      <tbody>{filas_html}</tbody>
    </table>
    <p style="color:gray;font-size:11px">Enviado automáticamente por Monitor de Impresoras — {datetime.now().strftime("%d/%m/%Y %H:%M")}</p>
    </body></html>
    """

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"[Impresoras] Alerta nivel bajo — {len(impresoras_bajas)} equipo(s) afectado(s)"
    msg["From"]    = remitente
    msg["To"]      = ", ".join(destinatarios)
    msg.attach(MIMEText(html, "html"))

    try:
        with smtplib.SMTP(servidor, puerto, timeout=15) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.login(remitente, password)
            smtp.sendmail(remitente, destinatarios, msg.as_string())
    except Exception as e:
        _log.error("Error enviando email de alerta: %s", e)


def abrir_configuracion():
    """Ventana de configuración general con pestañas: BD, Email, Monitoreo."""
    cfg = cargar_config()
    win = tk.Toplevel()
    win.title("Configuración")
    win.geometry("520x480")
    win.resizable(True, True)
    win.grab_set()

    # ── Notebook ──────────────────────────────────────────────────────────────
    notebook = ttk.Notebook(win)
    notebook.pack(fill="both", expand=True, padx=10, pady=(10, 4))

    lbl_kw = {"bg": BG_MAIN, "font": FONT_UI, "fg": "#333333"}

    # ═══════════════════════════════════════════════════════════════════════════
    # Pestaña 1: Base de datos
    # ═══════════════════════════════════════════════════════════════════════════
    frame_bd = ttk.Frame(notebook)
    notebook.add(frame_bd, text=" Base de datos ")

    nf_bd = tk.Frame(frame_bd, bg=BG_MAIN, padx=12, pady=12)
    nf_bd.pack(fill="both", expand=True)

    tk.Label(nf_bd, text="Archivo de base de datos compartida:", **lbl_kw).pack(anchor="w")

    var_db = tk.StringVar(value=cfg.get("db_path", "") or _DB_PATH_DEFAULT)
    f_db_row = tk.Frame(nf_bd, bg=BG_MAIN)
    f_db_row.pack(fill="x", pady=(4, 6))
    f_db_row.columnconfigure(0, weight=1)
    entry_db = tk.Entry(f_db_row, textvariable=var_db, font=("Segoe UI", 8))
    entry_db.grid(row=0, column=0, sticky="ew", padx=(0, 6))

    def _examinar_db():
        ruta = filedialog.asksaveasfilename(
            title="Seleccionar archivo de base de datos",
            defaultextension=".db", filetypes=[("SQLite DB", "*.db"), ("Todos", "*.*")],
            initialfile=os.path.basename(var_db.get()),
            initialdir=os.path.dirname(var_db.get()), parent=win)
        if ruta:
            var_db.set(ruta.replace("/", "\\"))

    btn_exam = tk.Button(f_db_row, text="Examinar\u2026", command=_examinar_db)
    _estilo_btn(btn_exam, primario=False)
    btn_exam.grid(row=0, column=1)

    def _usar_default():
        var_db.set(_DB_PATH_DEFAULT)

    tk.Label(nf_bd, text=f"Predeterminada: {_DB_PATH_DEFAULT}",
             bg=BG_MAIN, font=("Segoe UI", 7), fg="#888888", wraplength=460,
             justify="left").pack(anchor="w")

    btn_reset_db = tk.Button(nf_bd, text="Usar predeterminada", command=_usar_default)
    _estilo_btn(btn_reset_db, primario=False)
    btn_reset_db.pack(anchor="w", pady=(4, 0))

    lbl_db_status = tk.Label(nf_bd, text="", bg=BG_MAIN, font=("Segoe UI", 8))
    lbl_db_status.pack(anchor="w", pady=(6, 0))

    def _probar_db():
        ruta = var_db.get().strip()
        if os.path.isfile(ruta):
            try:
                with sqlite3.connect(ruta, timeout=5) as conn:
                    conn.execute("SELECT 1")
                lbl_db_status.config(text="\u2713 Conexión exitosa", fg="green")
            except Exception as e:
                lbl_db_status.config(text=f"\u2717 Error: {e}", fg="red")
        else:
            lbl_db_status.config(text="\u26a0 El archivo no existe (se creará al iniciar)", fg="#E65100")

    btn_probar_db = tk.Button(nf_bd, text="Probar conexión", command=_probar_db)
    _estilo_btn(btn_probar_db, primario=False)
    btn_probar_db.pack(anchor="w", pady=(2, 0))

    tk.Label(nf_bd,
             text="\u2139 Para compartir la BD entre PCs, apuntá a una carpeta de red.\n"
                  "  Ej: \\\\servidor\\share\\impresoras.db\n"
                  "  Los cambios de ruta se aplican al reiniciar la aplicación.",
             bg=BG_MAIN, font=("Segoe UI", 7), fg="#1565C0",
             wraplength=460, justify="left").pack(anchor="w", pady=(10, 0))

    # ═══════════════════════════════════════════════════════════════════════════
    # Pestaña 2: Correo electrónico
    # ═══════════════════════════════════════════════════════════════════════════
    frame_mail = ttk.Frame(notebook)
    notebook.add(frame_mail, text=" Correo electrónico ")

    nf_mail = tk.Frame(frame_mail, bg=BG_MAIN, padx=12, pady=12)
    nf_mail.pack(fill="both", expand=True)

    var_hab       = tk.BooleanVar(value=bool(cfg.get("email_habilitado", False)))
    var_remitente = tk.StringVar(value=cfg.get("email_remitente", ""))
    var_password  = tk.StringVar(value=cfg.get("email_password", ""))
    var_dest      = tk.StringVar(value=cfg.get("email_destinatarios", ""))
    var_servidor  = tk.StringVar(value=cfg.get("email_servidor", "smtp.office365.com"))
    var_puerto    = tk.StringVar(value=str(cfg.get("email_puerto", 587)))

    chk_hab = tk.Checkbutton(nf_mail, text="Habilitar notificaciones por email",
                             variable=var_hab, bg=BG_MAIN, font=FONT_UI,
                             activebackground=BG_MAIN)
    chk_hab.pack(anchor="w", pady=(0, 8))

    f_mail_campos = tk.Frame(nf_mail, bg=BG_MAIN)
    f_mail_campos.pack(fill="x")
    f_mail_campos.columnconfigure(1, weight=1)

    mail_campos = [
        ("Remitente:",          var_remitente, False, "Correo electrónico que enviará las alertas"),
        ("Contraseña:",         var_password,  True,  "Contraseña del remitente (se guarda en texto plano)"),
        ("Destinatarios:",      var_dest,      False, "Separar múltiples destinatarios con coma"),
        ("Servidor SMTP:",      var_servidor,  False, "Ej: smtp.office365.com, smtp.gmail.com"),
        ("Puerto:",             var_puerto,    False, "Puerto SMTP con STARTTLS (587) o SSL (465)"),
    ]

    entries_mail = {}
    for i, (label, var, es_pass, tip) in enumerate(mail_campos):
        tk.Label(f_mail_campos, text=label, anchor="w", bg=BG_MAIN,
                 font=FONT_UI).grid(row=i, column=0, sticky="e", pady=3, padx=(0, 6))
        e = tk.Entry(f_mail_campos, textvariable=var, width=28, font=FONT_UI,
                     show="*" if es_pass else "")
        e.grid(row=i, column=1, sticky="ew", padx=(0, 4), pady=3)
        entries_mail[label] = e
        tk.Label(f_mail_campos, text=tip, bg=BG_MAIN, font=("Segoe UI", 7),
                 fg="#999999").grid(row=i, column=2, sticky="w", padx=(0, 0))

    # Toggle contraseña
    f_pass = tk.Frame(nf_mail, bg=BG_MAIN)
    f_pass.pack(fill="x", pady=(4, 0))
    tk.Button(f_pass, text="Ver / Ocultar contraseña",
              command=lambda: entries_mail["Contraseña:"].config(
                  show="" if entries_mail["Contraseña:"].cget("show") == "*" else "*"),
              font=FONT_UI, relief="flat", bg="#EEEEEE", cursor="hand2"
              ).pack(side="left")
    lbl_mail_status = tk.Label(f_pass, text="", bg=BG_MAIN, font=("Segoe UI", 8))
    lbl_mail_status.pack(side="left", padx=(10, 0))

    # ═══════════════════════════════════════════════════════════════════════════
    # Pestaña 3: Monitoreo
    # ═══════════════════════════════════════════════════════════════════════════
    frame_mon = ttk.Frame(notebook)
    notebook.add(frame_mon, text=" Monitoreo ")

    nf_mon = tk.Frame(frame_mon, bg=BG_MAIN, padx=12, pady=12)
    nf_mon.pack(fill="both", expand=True)

    var_umbral_bajo  = tk.IntVar(value=cfg.get("umbral_bajo", 10))
    var_umbral_medio = tk.IntVar(value=cfg.get("umbral_medio", 25))
    var_max_workers  = tk.IntVar(value=cfg.get("max_workers", 20))
    var_intervalo    = tk.StringVar(value=cfg.get("intervalo_auto", "1 hora"))

    tk.Label(nf_mon, text="Umbrales de alerta (porcentaje de vida útil):",
             **lbl_kw).pack(anchor="w", pady=(0, 8))

    f_umbrales = tk.Frame(nf_mon, bg=BG_MAIN)
    f_umbrales.pack(fill="x", pady=(0, 12))

    tk.Label(f_umbrales, text="Nivel bajo:", bg=BG_MAIN, font=FONT_UI).pack(side="left")
    s_bajo = Spinbox(f_umbrales, from_=1, to=99, width=4, textvariable=var_umbral_bajo, font=FONT_UI)
    s_bajo.pack(side="left", padx=(4, 6))
    tk.Label(f_umbrales, text="%", bg=BG_MAIN, font=FONT_UI).pack(side="left", padx=(0, 10))

    tk.Label(f_umbrales, text="Nivel medio:", bg=BG_MAIN, font=FONT_UI).pack(side="left")
    s_medio = Spinbox(f_umbrales, from_=1, to=99, width=4, textvariable=var_umbral_medio, font=FONT_UI)
    s_medio.pack(side="left", padx=(4, 6))
    tk.Label(f_umbrales, text="%", bg=BG_MAIN, font=FONT_UI).pack(side="left")

    # Separador
    tk.Frame(nf_mon, bg="#DDDDDD", height=1, bd=0, highlightthickness=0).pack(
        fill="x", pady=(0, 10))

    tk.Label(nf_mon, text="Rendimiento:", **lbl_kw).pack(anchor="w", pady=(0, 6))

    f_workers = tk.Frame(nf_mon, bg=BG_MAIN)
    f_workers.pack(fill="x", pady=(0, 8))
    tk.Label(f_workers, text="Máx. hilos simultáneos:", bg=BG_MAIN, font=FONT_UI).pack(side="left")
    s_workers = Spinbox(f_workers, from_=1, to=100, width=4, textvariable=var_max_workers, font=FONT_UI)
    s_workers.pack(side="left", padx=(6, 0))
    tk.Label(f_workers, text="  (más hilos = más velocidad, más CPU)",
             bg=BG_MAIN, font=("Segoe UI", 7), fg="#999999").pack(side="left")

    # Separador
    tk.Frame(nf_mon, bg="#DDDDDD", height=1, bd=0, highlightthickness=0).pack(
        fill="x", pady=(0, 10))

    tk.Label(nf_mon, text="Monitoreo automático:", **lbl_kw).pack(anchor="w", pady=(0, 6))

    f_intervalo = tk.Frame(nf_mon, bg=BG_MAIN)
    f_intervalo.pack(fill="x")
    tk.Label(f_intervalo, text="Intervalo predeterminado:", bg=BG_MAIN, font=FONT_UI).pack(side="left")
    combo_int = Combobox(f_intervalo, textvariable=var_intervalo,
                         values=list(INTERVALOS_AUTO.keys()), state="readonly", width=10)
    combo_int.pack(side="left", padx=(6, 0))
    tk.Label(f_intervalo, text="  (se aplica al reiniciar)",
             bg=BG_MAIN, font=("Segoe UI", 7), fg="#999999").pack(side="left")

    # ═══════════════════════════════════════════════════════════════════════════
    # Botones inferiores
    # ═══════════════════════════════════════════════════════════════════════════
    frame_btns = tk.Frame(win, bg=BG_MAIN)
    frame_btns.pack(fill="x", padx=14, pady=(4, 10))

    def guardar():
        nueva_ruta = var_db.get().strip()
        try:
            puerto = int(var_puerto.get().strip() or 587)
        except ValueError:
            puerto = 587
        guardar_config(
            db_path             = nueva_ruta if nueva_ruta != _DB_PATH_DEFAULT else "",
            email_habilitado    = var_hab.get(),
            email_remitente     = var_remitente.get().strip(),
            email_password      = var_password.get(),
            email_destinatarios = var_dest.get().strip(),
            email_servidor      = var_servidor.get().strip(),
            email_puerto        = puerto,
            umbral_bajo         = var_umbral_bajo.get(),
            umbral_medio        = var_umbral_medio.get(),
            max_workers         = var_max_workers.get(),
            intervalo_auto      = var_intervalo.get(),
        )
        _inicializar_db_path()
        messagebox.showinfo(
            "Guardado",
            "Configuración guardada.\n\n"
            "La nueva ruta de base de datos, umbrales e intervalo\n"
            "se aplicarán en el próximo inicio de la aplicación.",
            parent=win)
        win.destroy()

    def probar_envio():
        guardar_config(
            email_habilitado    = var_hab.get(),
            email_remitente     = var_remitente.get().strip(),
            email_password      = var_password.get(),
            email_destinatarios = var_dest.get().strip(),
            email_servidor      = var_servidor.get().strip(),
            email_puerto        = int(var_puerto.get().strip() or 587),
        )

        def _enviar():
            cfg2  = cargar_config()
            rem   = cfg2.get("email_remitente", "").strip()
            pwd   = cfg2.get("email_password", "")
            dest_raw = cfg2.get("email_destinatarios", "").strip()
            srv   = cfg2.get("email_servidor", "smtp.office365.com")
            prt   = int(cfg2.get("email_puerto", 587))
            dests = [d.strip() for d in dest_raw.split(",") if d.strip()]
            if not rem or not dests:
                win.after(0, lambda: messagebox.showerror(
                    "Error", "Complete remitente y destinatarios antes de probar.", parent=win))
                return
            lbl_mail_status.config(text="Enviando\u2026", fg="#555555")
            msg = MIMEMultipart("alternative")
            msg["Subject"] = "Prueba de configuración \u2014 Monitor de Impresoras"
            msg["From"] = rem; msg["To"] = ", ".join(dests)
            msg.attach(MIMEText(
                "<h3>Prueba exitosa</h3><p>La configuración de email funciona correctamente.</p>", "html"))
            try:
                with smtplib.SMTP(srv, prt, timeout=15) as smtp:
                    smtp.ehlo(); smtp.starttls(); smtp.login(rem, pwd)
                    smtp.sendmail(rem, dests, msg.as_string())
                win.after(0, lambda: lbl_mail_status.config(text="\u2713 Enviado", fg="green"))
                win.after(0, lambda: messagebox.showinfo(
                    "Prueba exitosa", "El correo de prueba fue enviado correctamente.", parent=win))
            except Exception as exc:
                win.after(0, lambda: lbl_mail_status.config(text="\u2717 Error", fg="red"))
                win.after(0, messagebox.showerror, "Error al enviar", str(exc))

        threading.Thread(target=_enviar, daemon=True).start()

    btn_guardar = tk.Button(frame_btns, text="Guardar", command=guardar, width=12)
    _estilo_btn(btn_guardar, primario=True)
    btn_guardar.pack(side="left", padx=6)

    btn_probar = tk.Button(frame_btns, text="Probar envío", command=probar_envio, width=12)
    _estilo_btn(btn_probar, primario=False)
    btn_probar.pack(side="left", padx=6)

    btn_cancelar = tk.Button(frame_btns, text="Cancelar", command=win.destroy, width=10)
    _estilo_btn(btn_cancelar, primario=False)
    btn_cancelar.pack(side="left", padx=6)
# ---------------------------------------------------------------------------
# Exportar a Excel
# ---------------------------------------------------------------------------

def exportar_excel(ctx):
    """Exporta la tabla actual (ctx.filas_tabla) a un .xlsx con 2 hojas."""
    if not ctx.filas_tabla:
        messagebox.showwarning("Sin datos", "No hay datos en la tabla para exportar.\n"
                               "Ejecute un monitoreo primero.")
        return

    nombre_default = f"monitoreo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    ruta = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile=nombre_default,
        title="Guardar exportación",
    )
    if not ruta:
        return

    wb = Workbook()

    # ── Hoja 1: detalle del monitoreo ─────────────────────────────────────
    ws = wb.active
    ws.title = "Monitoreo"

    fill_header    = PatternFill("solid", fgColor="CCCCCC")
    fill_bajo      = PatternFill("solid", fgColor="FF9999")
    fill_medio     = PatternFill("solid", fgColor="FFFF99")
    fill_sin_datos = PatternFill("solid", fgColor="D0D0D0")
    font_bold      = Font(bold=True)
    align_center   = Alignment(horizontal="center")

    nivel_texto = {"bajo": "Bajo", "medio": "Medio", "sin_datos": "Sin datos", "": "OK"}
    nivel_fill  = {"bajo": fill_bajo, "medio": fill_medio,
                   "sin_datos": fill_sin_datos, "": None}

    cabeceras = ["Sucursal", "IP", "Modelo",
                 "Tóner (%)", "Unidad Imagen (%)", "Kit Mantenimiento (%)", "Nivel"]
    ws.append(cabeceras)
    for cell in ws[1]:
        cell.font      = font_bold
        cell.fill      = fill_header
        cell.alignment = align_center

    for fila in ctx.filas_tabla:
        tag  = fila[7]
        row  = list(fila[:7]) + [nivel_texto.get(tag, tag)]
        ws.append(row)
        fill = nivel_fill.get(tag)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    # Ajustar anchos
    anchos = [16, 14, 22, 12, 18, 20, 10]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = ancho

    # ── Hoja 2: resumen por sucursal ──────────────────────────────────────
    ws2 = wb.create_sheet("Por Sucursal")
    cab2 = ["Sucursal", "Total", "OK", "Nivel medio", "Nivel bajo", "Sin datos"]
    ws2.append(cab2)
    for cell in ws2[1]:
        cell.font      = font_bold
        cell.fill      = fill_header
        cell.alignment = align_center

    conteo = {}
    for fila in ctx.filas_tabla:
        suc = fila[0] or "(sin sucursal)"
        tag = fila[6]
        if suc not in conteo:
            conteo[suc] = {"total": 0, "": 0, "medio": 0, "bajo": 0, "sin_datos": 0}
        conteo[suc]["total"] += 1
        conteo[suc][tag]     += 1

    # Ordenar: mayor nivel bajo primero, luego nivel medio
    filas_suc = sorted(conteo.items(),
                       key=lambda x: (-x[1]["bajo"], -x[1]["medio"]))
    for suc, c in filas_suc:
        row2 = [suc, c["total"], c[""], c["medio"], c["bajo"], c["sin_datos"]]
        ws2.append(row2)
        r = ws2.max_row
        if c["bajo"]      > 0: ws2.cell(r, 5).fill = fill_bajo
        if c["medio"]     > 0: ws2.cell(r, 4).fill = fill_medio
        if c["sin_datos"] > 0: ws2.cell(r, 6).fill = fill_sin_datos

    for i, ancho in enumerate([20, 8, 8, 12, 12, 12], 1):
        ws2.column_dimensions[ws2.cell(1, i).column_letter].width = ancho

    try:
        wb.save(ruta)
        messagebox.showinfo("Exportado", f"Archivo guardado:\n{ruta}", parent=ctx.ventana)
    except Exception as e:
        messagebox.showerror("Error al guardar", str(e), parent=ctx.ventana)


# ---------------------------------------------------------------------------
# Gráfico de tendencia + predicción
# ---------------------------------------------------------------------------

def mostrar_grafico(ip):
    """Abre una ventana Toplevel con gráfico embebido en tkinter y controles interactivos."""
    # ── Carga de datos ──────────────────────────────────────────────────────
    with db_connect() as conn:
        rows = conn.execute(
            "SELECT fecha, toner, unidad_imagen, kit_mantenimiento "
            "FROM monitoreos WHERE ip=? ORDER BY fecha",
            (ip,),
        ).fetchall()

    if not rows:
        messagebox.showinfo("Sin datos", f"No hay historial para la IP {ip}.")
        return

    all_fechas, all_toner, all_unidad, all_kit = [], [], [], []
    for r in rows:
        try:
            fecha_dt = datetime.strptime(r[0], "%Y-%m-%d %H:%M:%S")
        except (ValueError, TypeError):
            continue
        all_fechas.append(fecha_dt)
        all_toner.append(r[1] * 100 if r[1] is not None else float('nan'))
        all_unidad.append(r[2] * 100 if r[2] is not None else float('nan'))
        all_kit.append(r[3]   * 100 if r[3] is not None else float('nan'))

    if not all_fechas:
        messagebox.showinfo("Sin datos", f"No hay historial válido para la IP {ip}.")
        return

    # Título: obtener modelo/sucursal del catálogo
    titulo = ip
    todas = db_impresoras_todas()
    for imp in todas:
        if imp["ip"] == ip:
            modelo   = imp["modelo"]   or ""
            sucursal = imp["sucursal"] or ""
            titulo   = f"{ip} — {modelo} | {sucursal}" if modelo else ip
            break

    # ── Ventana Toplevel ─────────────────────────────────────────────────────
    win = tk.Toplevel()
    win.title(f"Gráfico: {titulo}")
    win.geometry("920x600")
    win.minsize(700, 480)
    win.config(bg=BG_MAIN)
    win.columnconfigure(0, weight=1)
    win.rowconfigure(2, weight=1)

    # ── Figura matplotlib ─────────────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(9, 4.5))
    fig.patch.set_facecolor("white")

    canvas = FigureCanvasTkAgg(fig, master=win)
    canvas.get_tk_widget().grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 0))

    # ── Variables de series ───────────────────────────────────────────────────
    var_toner  = tk.BooleanVar(value=True)
    var_unidad = tk.BooleanVar(value=True)
    var_kit    = tk.BooleanVar(value=True)

    # ── Función de dibujo ─────────────────────────────────────────────────────
    lbl_pred = tk.Label(win, text="", bg=BG_MAIN, font=("Segoe UI", 8),
                        fg="#555555", anchor="w")

    def _dibujar(desde_dt=None, hasta_dt=None):
        ax.clear()

        pares = list(zip(all_fechas, all_toner, all_unidad, all_kit))
        if desde_dt:
            pares = [(f, t, u, k) for f, t, u, k in pares if f >= desde_dt]
        if hasta_dt:
            pares = [(f, t, u, k) for f, t, u, k in pares if f <= hasta_dt]

        if not pares:
            ax.text(0.5, 0.5, "Sin datos en el rango seleccionado",
                    ha="center", va="center", transform=ax.transAxes,
                    fontsize=11, color="#888888")
            canvas.draw()
            lbl_pred.config(text="")
            return

        fechas, toner_v, unidad_v, kit_v = zip(*pares)

        # Estilo área
        ax.set_facecolor("#FAFAFA")
        ax.grid(True, alpha=0.25, linestyle="--", color="#AAAAAA")
        ax.spines[["top", "right"]].set_visible(False)
        ax.spines[["left", "bottom"]].set_color("#CCCCCC")
        ax.set_title(titulo, fontsize=10, color="#333333", pad=8)

        mostrar_labels = len(fechas) <= 20

        series_def = [
            ("Tóner",         toner_v,   "#2196F3", var_toner),
            ("Unidad Imagen", unidad_v,  "#FF9800", var_unidad),
            ("Kit Mant.",     kit_v,     "#4CAF50", var_kit),
        ]
        pred_textos = []
        for label, vals, color, var in series_def:
            if not var.get():
                continue
            vals_list = list(vals)
            if not any(v == v for v in vals_list):
                continue

            ax.plot(fechas, vals_list, marker='o', markersize=5,
                    linewidth=2, label=label, color=color, zorder=3)

            if mostrar_labels:
                for f, v in zip(fechas, vals_list):
                    if v == v:  # no NaN
                        ax.annotate(f"{v:.0f}%", (f, v),
                                    textcoords="offset points", xytext=(0, 8),
                                    ha="center", fontsize=7, color=color,
                                    fontweight="bold")

            fecha_pred = predecir_agotamiento(list(fechas), vals_list)
            if fecha_pred:
                ax.axvline(x=fecha_pred, color=color, linestyle=":", alpha=0.7,
                           linewidth=1.5,
                           label=f"Pred. {label[:3]}: {fecha_pred.strftime('%d/%m/%Y')}")
                pred_textos.append(f"{label}: {fecha_pred.strftime('%d/%m/%Y')}")
            else:
                pred_textos.append(f"{label}: —")

        # Umbrales
        cfg = cargar_config()
        ubajo  = cfg.get("umbral_bajo",  10)
        umedio = cfg.get("umbral_medio", 25)
        ax.axhline(y=ubajo,  color="#E53935", linestyle="--", alpha=0.4,
                   linewidth=1, label=f"Umbral bajo ({ubajo}%)")
        ax.axhline(y=umedio, color="#FB8C00", linestyle="--", alpha=0.4,
                   linewidth=1, label=f"Umbral medio ({umedio}%)")

        # Formato ejes
        ax.set_ylim(0, 110)
        ax.yaxis.set_major_formatter(
            mticker.FuncFormatter(lambda x, _: f"{x:.0f}%"))
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%d/%m/%y"))
        fig.autofmt_xdate(rotation=35)
        ax.set_ylabel("Nivel (%)", fontsize=9, color="#555555")
        ax.legend(loc="best", fontsize=8, framealpha=0.7)

        fig.tight_layout(pad=1.5)
        canvas.draw()

        if pred_textos:
            lbl_pred.config(
                text="Predicción de agotamiento:  " + "   |   ".join(pred_textos))
        else:
            lbl_pred.config(text="")

    # ── Redibujar leyendo entries de fecha ────────────────────────────────────
    def _redibujar(*_):
        desde_dt = hasta_dt = None
        txt_d = entry_desde.get().strip()
        txt_h = entry_hasta.get().strip()
        try:
            if txt_d:
                desde_dt = datetime.strptime(txt_d, "%d/%m/%Y")
        except ValueError:
            pass
        try:
            if txt_h:
                hasta_dt = datetime.strptime(txt_h, "%d/%m/%Y").replace(
                    hour=23, minute=59, second=59)
        except ValueError:
            pass
        _dibujar(desde_dt, hasta_dt)

    def _aplicar():
        _redibujar()

    def _todo():
        entry_desde.delete(0, tk.END)
        entry_hasta.delete(0, tk.END)
        _dibujar()

    def _exportar():
        ruta = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("PNG", "*.png"), ("JPEG", "*.jpg")],
            initialfile=f"grafico_{ip}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png",
            parent=win,
        )
        if ruta:
            fig.savefig(ruta, dpi=150, bbox_inches="tight")
            messagebox.showinfo("Exportado", f"Imagen guardada:\n{ruta}", parent=win)

    # ── Row 0: Checkboxes + Exportar PNG ─────────────────────────────────────
    frame_top = tk.Frame(win, bg=BG_MAIN)
    frame_top.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 4))
    frame_top.columnconfigure(3, weight=1)

    chk_style = {"bg": BG_MAIN, "font": ("Segoe UI", 9), "fg": "#333333",
                 "activebackground": BG_MAIN, "cursor": "hand2"}
    tk.Checkbutton(frame_top, text="Tóner",         variable=var_toner,
                   **chk_style).grid(row=0, column=0, padx=(0, 8))
    tk.Checkbutton(frame_top, text="Unidad Imagen", variable=var_unidad,
                   **chk_style).grid(row=0, column=1, padx=(0, 8))
    tk.Checkbutton(frame_top, text="Kit Mant.",     variable=var_kit,
                   **chk_style).grid(row=0, column=2, padx=(0, 8))

    btn_exportar = tk.Button(frame_top, text="Exportar PNG", command=_exportar)
    _estilo_btn(btn_exportar, primario=False)
    btn_exportar.grid(row=0, column=4, sticky="e")

    var_toner.trace_add("write",  lambda *_: _redibujar())
    var_unidad.trace_add("write", lambda *_: _redibujar())
    var_kit.trace_add("write",    lambda *_: _redibujar())

    # ── Row 1: Filtro de fechas ───────────────────────────────────────────────
    frame_filtro = tk.Frame(win, bg=BG_MAIN)
    frame_filtro.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 6))

    lbl_kw = {"bg": BG_MAIN, "font": ("Segoe UI", 9), "fg": "#555555"}
    tk.Label(frame_filtro, text="Desde:", **lbl_kw).grid(row=0, column=0, padx=(0, 4))
    entry_desde = tk.Entry(frame_filtro, width=12, font=("Segoe UI", 9))
    entry_desde.grid(row=0, column=1, padx=(0, 12))

    tk.Label(frame_filtro, text="Hasta:", **lbl_kw).grid(row=0, column=2, padx=(0, 4))
    entry_hasta = tk.Entry(frame_filtro, width=12, font=("Segoe UI", 9))
    entry_hasta.grid(row=0, column=3, padx=(0, 12))

    btn_aplicar = tk.Button(frame_filtro, text="Aplicar", command=_aplicar)
    _estilo_btn(btn_aplicar, primario=True)
    btn_aplicar.grid(row=0, column=4, padx=(0, 6))

    btn_todo = tk.Button(frame_filtro, text="Todo", command=_todo)
    _estilo_btn(btn_todo, primario=False)
    btn_todo.grid(row=0, column=5)

    tk.Label(frame_filtro, text="(DD/MM/YYYY)", bg=BG_MAIN,
             font=("Segoe UI", 7), fg="#AAAAAA").grid(row=0, column=6, padx=(8, 0))

    # ── Row 3: NavigationToolbar ──────────────────────────────────────────────
    frame_tb = tk.Frame(win, bg=BG_MAIN)
    frame_tb.grid(row=3, column=0, sticky="ew", padx=10)
    NavigationToolbar2Tk(canvas, frame_tb).update()

    # ── Row 4: Panel de predicciones ─────────────────────────────────────────
    lbl_pred.grid(row=4, column=0, sticky="ew", padx=12, pady=(2, 8))

    # ── Cierre seguro ─────────────────────────────────────────────────────────
    win.protocol("WM_DELETE_WINDOW", lambda: (plt.close(fig), win.destroy()))

    # ── Rango predeterminado: últimos 30 días ─────────────────────────────────
    _hoy   = datetime.now()
    _desde = _hoy - timedelta(days=30)
    entry_desde.insert(0, _desde.strftime("%d/%m/%Y"))
    entry_hasta.insert(0, _hoy.strftime("%d/%m/%Y"))

    # ── Dibujo inicial con el filtro ya activo ────────────────────────────────
    _redibujar()

# ---------------------------------------------------------------------------
# Orquestador principal  (corre en hilo secundario)
# ---------------------------------------------------------------------------

def ejecutar_monitoreo(ctx, umbral_bajo, umbral_medio, es_automatico=False):
    """Consulta todas las impresoras activas en paralelo y guarda resultados en la DB."""
    try:
        impresoras = db_impresoras_todas(activas_solo=True)

        if not impresoras:
            ctx.ventana.after(0, _finalizar, ctx,
                              "No hay impresoras activas en la base de datos.\n"
                              "Agregue impresoras usando el botón 'Impresoras'.", "error")
            return

        total_impresoras = len(impresoras)
        fecha_actual     = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        resultados       = {}
        completados      = 0

        with ThreadPoolExecutor(max_workers=cargar_config().get("max_workers", 20)) as executor:
            future_to_info = {
                executor.submit(obtener_status, imp["ip"], imp["modelo"], ctx.evento_cancelar):
                    (imp["ip"], imp["modelo"], imp["sucursal"])
                for imp in impresoras
            }

            for future in as_completed(future_to_info):
                if ctx.evento_cancelar.is_set():
                    break

                ip, modelo, sucursal = future_to_info[future]
                toner, kit, unidad   = future.result()
                resultados[ip]       = (ip, modelo, sucursal, toner, kit, unidad)

                completados += 1
                progreso     = int((completados / total_impresoras) * 100)

                ctx.ventana.after(0, _actualizar_progreso,
                                  ctx, sucursal, ip, modelo, toner, kit, unidad,
                                  progreso, umbral_bajo, umbral_medio, fecha_actual)

        if ctx.evento_cancelar.is_set():
            ctx.ventana.after(0, _finalizar, ctx, "Monitoreo cancelado.", "cancelado")
            return

        # Guardar resultados en la DB
        with db_connect() as conn:
            for ip, modelo, sucursal, toner, kit, unidad in resultados.values():
                conn.execute(
                    "INSERT INTO monitoreos (fecha, ip, toner, unidad_imagen, kit_mantenimiento) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (fecha_actual, ip, toner, unidad, kit),
                )

        # Calcular estadísticas
        n_bajo = n_medio = n_sin_datos = 0
        for ip, modelo, sucursal, toner, kit, unidad in resultados.values():
            valores = [v for v in (toner, kit, unidad) if v is not None]
            nivel   = clasificar_nivel(valores, umbral_bajo, umbral_medio)
            if   nivel == "sin_datos": n_sin_datos += 1
            elif nivel == "bajo":      n_bajo      += 1
            elif nivel == "medio":     n_medio     += 1

        total        = len(resultados)
        respondieron = total - n_sin_datos
        ctx.ventana.after(0, _actualizar_resumen, ctx, total, respondieron, n_sin_datos, n_bajo, n_medio, fecha_actual)

        mensaje = f"Monitoreo completado. {total} impresoras consultadas."
        tipo    = "exito"
        if n_bajo > 0 or n_medio > 0 or n_sin_datos > 0:
            mensaje += (
                f"\n\nResumen de alertas:"
                f"\n  Nivel bajo (<{umbral_bajo}%): {n_bajo}"
                f"\n  Nivel medio (<{umbral_medio}%): {n_medio}"
                f"\n  Sin datos: {n_sin_datos}"
            )
            if n_bajo > 0:
                tipo = "alerta"

        # Enviar email si es monitoreo automático y hay nivel bajo
        if es_automatico and n_bajo > 0 and cargar_config().get("email_habilitado"):
            bajas = [
                (ip, m, s, t, k, u)
                for ip, m, s, t, k, u in resultados.values()
                if clasificar_nivel(
                    [v for v in (t, k, u) if v is not None], umbral_bajo, umbral_medio
                ) == "bajo"
            ]
            threading.Thread(
                target=enviar_alerta_email, args=(bajas, umbral_bajo), daemon=True
            ).start()

        ctx.ventana.after(0, _finalizar, ctx, mensaje, tipo)

    except Exception as e:
        ctx.ventana.after(0, _finalizar, ctx, f"Error: {e}", "error")

# ---------------------------------------------------------------------------
# Acciones de la UI
# ---------------------------------------------------------------------------

def iniciar_monitoreo(ctx, es_automatico=False):
    """Valida umbrales y lanza el monitoreo en un hilo separado."""
    try:
        umbral_bajo  = int(ctx.spinbox_bajo.get())
        umbral_medio = int(ctx.spinbox_medio.get())
    except ValueError:
        messagebox.showwarning("Umbrales inválidos",
                               "Los umbrales deben ser números enteros entre 1 y 99.")
        return

    if umbral_bajo >= umbral_medio:
        messagebox.showwarning("Umbrales inválidos",
                               f"El nivel bajo ({umbral_bajo}%) debe ser menor "
                               f"que el nivel medio ({umbral_medio}%).")
        return

    guardar_config(umbral_bajo=umbral_bajo, umbral_medio=umbral_medio)

    ctx.evento_cancelar.clear()
    ctx.filas_tabla.clear()
    ctx.sort_col = None
    ctx.sort_asc = True
    ctx.tree.delete(*ctx.tree.get_children())
    ctx.barra_progreso["value"] = 0
    ctx.btn_iniciar.config(state="disabled")
    ctx.btn_cancelar.config(state="normal")
    ctx.btn_grafico.config(state="disabled")
    ctx.btn_exportar.config(state="disabled")

    for col in COLUMNAS_TREE:
        ctx.tree.heading(col, text=col)

    hilo = threading.Thread(
        target=ejecutar_monitoreo,
        args=(ctx, umbral_bajo, umbral_medio, es_automatico),
        daemon=True,
    )
    hilo.start()


def ver_grafico(ctx):
    """Muestra el gráfico de tendencia para la impresora seleccionada en la tabla."""
    sel = ctx.tree.selection()
    if sel:
        ip = ctx.tree.item(sel[0], "values")[1]   # índice 1 = IP (sucursal en índice 0)
        mostrar_grafico(ip)

# ---------------------------------------------------------------------------
# Estadísticas de consumo por sucursal
# ---------------------------------------------------------------------------

def abrir_estadisticas_consumo():
    """Abre ventana con gráfico de barras, tabla resumen y detalle de consumo por sucursal."""
    win = tk.Toplevel()
    win.title("Estadísticas de Consumo por Sucursal")
    win.geometry("960x740")
    win.resizable(True, True)
    win.config(bg=BG_MAIN)
    win.columnconfigure(0, weight=1)
    win.rowconfigure(3, weight=1)

    hoy = datetime.now()
    lbl_kw = {"bg": BG_MAIN, "font": FONT_UI, "fg": "#555555"}
    suc_data_cache = {}
    suc_list_all = ["Todas"] + db_sucursales_activas()

    # ── Filtros ───────────────────────────────────────────────────────────────
    frame_f = tk.Frame(win, bg=BG_MAIN)
    frame_f.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 4))

    tk.Label(frame_f, text="Desde:", **lbl_kw).pack(side="left", padx=(0, 4))
    entry_desde = tk.Entry(frame_f, width=11, font=FONT_UI)
    entry_desde.insert(0, (hoy - timedelta(days=90)).strftime("%d/%m/%Y"))
    entry_desde.pack(side="left", padx=(0, 8))

    tk.Label(frame_f, text="Hasta:", **lbl_kw).pack(side="left", padx=(0, 4))
    entry_hasta = tk.Entry(frame_f, width=11, font=FONT_UI)
    entry_hasta.insert(0, hoy.strftime("%d/%m/%Y"))
    entry_hasta.pack(side="left", padx=(0, 8))

    tk.Label(frame_f, text="Tipo:", **lbl_kw).pack(side="left", padx=(0, 4))
    var_tipo = tk.StringVar(value="Todos")
    Combobox(frame_f, textvariable=var_tipo, values=["Todos"] + TIPOS_INSUMO,
             state="readonly", width=13).pack(side="left", padx=(0, 8))

    tk.Label(frame_f, text="Sucursal:", **lbl_kw).pack(side="left", padx=(0, 4))
    var_suc = tk.StringVar(value="Todas")
    combo_suc = Combobox(frame_f, textvariable=var_suc, values=suc_list_all,
                         state="readonly", width=16)
    combo_suc.pack(side="left", padx=(0, 8))

    tk.Label(frame_f, text="(DD/MM/YYYY)", bg=BG_MAIN,
             font=("Segoe UI", 7), fg="#AAAAAA").pack(side="left", padx=(0, 6))

    btn_aplicar = tk.Button(frame_f, text="Aplicar")
    _estilo_btn(btn_aplicar, primario=True)
    btn_aplicar.pack(side="left", padx=(0, 4))

    def _todo():
        entry_desde.delete(0, tk.END)
        entry_hasta.delete(0, tk.END)
        var_tipo.set("Todos"); var_suc.set("Todas")
        _dibujar()

    btn_todo = tk.Button(frame_f, text="Todo")
    _estilo_btn(btn_todo, primario=False)
    btn_todo.pack(side="left", padx=(0, 8))

    # Períodos rápidos
    def _set_periodo(dias):
        entry_desde.delete(0, tk.END)
        entry_desde.insert(0, (hoy - timedelta(days=dias)).strftime("%d/%m/%Y"))
        entry_hasta.delete(0, tk.END)
        entry_hasta.insert(0, hoy.strftime("%d/%m/%Y"))
        _dibujar()

    for texto, dias in [("7d", 7), ("30d", 30), ("90d", 90), ("1a", 365)]:
        b = tk.Button(frame_f, text=texto, width=3, command=lambda d=dias: _set_periodo(d))
        _estilo_btn(b, primario=False)
        b.pack(side="left", padx=(2, 0))

    btn_exportar = tk.Button(frame_f, text="Exportar Excel")
    _estilo_btn(btn_exportar, primario=False)
    btn_exportar.pack(side="right")

    # ── Canvas ────────────────────────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(9.0, 3.8))
    fig.patch.set_facecolor("white")
    canvas = FigureCanvasTkAgg(fig, master=win)
    canvas.get_tk_widget().grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 4))

    # ── Tabla resumen ─────────────────────────────────────────────────────────
    frame_tbl = tk.Frame(win, bg=BG_MAIN)
    frame_tbl.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 2))

    cols_res = ("Sucursal", "Tóner", "Unidad Imagen", "Total")
    tree_res = Treeview(frame_tbl, columns=cols_res, show="headings", height=5)
    for col in cols_res:
        tree_res.heading(col, text=col)
        tree_res.column(col, anchor="center", width=280 if col == "Sucursal" else 130)
    sb_res = Scrollbar(frame_tbl, orient="vertical", command=tree_res.yview)
    tree_res.configure(yscrollcommand=sb_res.set)
    tree_res.pack(side="left", fill="x", expand=True)
    sb_res.pack(side="right", fill="y")

    # ── Panel detalle por sucursal ────────────────────────────────────────────
    frame_det = tk.LabelFrame(win, text="Detalle por sucursal", bg=BG_MAIN,
                              font=FONT_BOLD, padx=6, pady=4)
    frame_det.grid(row=3, column=0, sticky="nsew", padx=10, pady=(4, 8))
    frame_det.columnconfigure(0, weight=1)
    frame_det.rowconfigure(0, weight=1)

    cols_det = ("Fecha", "Tipo", "Modelo", "Cantidad")
    tree_det = Treeview(frame_det, columns=cols_det, show="headings", height=5)
    for col in cols_det:
        tree_det.heading(col, text=col)
        tree_det.column(col, anchor="center", width=160 if col == "Modelo" else 100)
    sb_det = Scrollbar(frame_det, orient="vertical", command=tree_det.yview)
    tree_det.configure(yscrollcommand=sb_det.set)
    tree_det.pack(side="left", fill="both", expand=True)
    sb_det.pack(side="right", fill="y")

    # ── Funciones ────────────────────────────────────────────────────────────
    def _parsear_fechas():
        desde_dt = hasta_dt = None
        try:
            txt = entry_desde.get().strip()
            if txt: desde_dt = datetime.strptime(txt, "%d/%m/%Y")
        except ValueError: pass
        try:
            txt = entry_hasta.get().strip()
            if txt: hasta_dt = datetime.strptime(txt, "%d/%m/%Y")
        except ValueError: pass
        desde_str = desde_dt.strftime("%Y-%m-%d") if desde_dt else "1900-01-01"
        hasta_str = hasta_dt.strftime("%Y-%m-%d 23:59:59") if hasta_dt else "2099-12-31"
        return desde_str, hasta_str, desde_dt, hasta_dt

    def _dibujar():
        nonlocal suc_data_cache
        ax.clear()
        tree_res.delete(*tree_res.get_children())

        desde_str, hasta_str, desde_dt, hasta_dt = _parsear_fechas()
        tipo_f = var_tipo.get()
        suc_f  = var_suc.get()

        with db_connect() as conn:
            q = ("SELECT sucursal, tipo_insumo, SUM(cantidad) as total "
                 "FROM envios WHERE fecha BETWEEN ? AND ?")
            params = [desde_str, hasta_str]
            if tipo_f != "Todos":
                q += " AND tipo_insumo = ?"; params.append(tipo_f)
            if suc_f != "Todas":
                q += " AND sucursal = ?"; params.append(suc_f)
            q += " GROUP BY sucursal, tipo_insumo ORDER BY sucursal"
            rows = conn.execute(q, params).fetchall()

        if not rows:
            ax.text(0.5, 0.5, "Sin datos para el período seleccionado",
                    ha="center", va="center", transform=ax.transAxes,
                    fontsize=11, color="#888888")
            canvas.draw()
            tree_res.insert("", "end", values=("(sin datos)", "—", "—", "—"))
            suc_data_cache = {}
            return

        suc_data_cache = {}
        for suc, tipo, cant in rows:
            if suc not in suc_data_cache:
                suc_data_cache[suc] = {"Tóner": 0, "Unidad Imagen": 0}
            suc_data_cache[suc][tipo] = suc_data_cache[suc].get(tipo, 0) + cant

        sucursales = list(suc_data_cache.keys())
        toner_vals  = [suc_data_cache[s].get("Tóner", 0) for s in sucursales]
        unidad_vals = [suc_data_cache[s].get("Unidad Imagen", 0) for s in sucursales]

        y = np.arange(len(sucursales))
        h = 0.35
        mostrar_toner  = tipo_f in ("Todos", "Tóner")
        mostrar_unidad = tipo_f in ("Todos", "Unidad Imagen")

        if mostrar_toner and mostrar_unidad:
            bars1 = ax.barh(y + h / 2, toner_vals,  h, label="Tóner",
                            color="#2196F3", zorder=3)
            bars2 = ax.barh(y - h / 2, unidad_vals, h, label="Unidad Imagen",
                            color="#FF9800", zorder=3)
            for bar in list(bars1) + list(bars2):
                w = bar.get_width()
                if w > 0:
                    ax.text(w + 0.1, bar.get_y() + bar.get_height() / 2,
                            str(int(w)), va="center", ha="left", fontsize=8)
        elif mostrar_toner:
            bars = ax.barh(y, toner_vals, 0.55, label="Tóner", color="#2196F3", zorder=3)
            for bar in bars:
                w = bar.get_width()
                if w > 0:
                    ax.text(w + 0.1, bar.get_y() + bar.get_height() / 2,
                            str(int(w)), va="center", ha="left", fontsize=8)
        else:
            bars = ax.barh(y, unidad_vals, 0.55, label="Unidad Imagen",
                           color="#FF9800", zorder=3)
            for bar in bars:
                w = bar.get_width()
                if w > 0:
                    ax.text(w + 0.1, bar.get_y() + bar.get_height() / 2,
                            str(int(w)), va="center", ha="left", fontsize=8)

        ax.set_yticks(y)
        ax.set_yticklabels(sucursales, fontsize=8)
        ax.set_xlabel("Unidades enviadas", fontsize=9)
        ax.set_facecolor("#FAFAFA")
        ax.grid(True, axis="x", alpha=0.25, linestyle="--", color="#AAAAAA")
        ax.spines[["top", "right"]].set_visible(False)
        ax.spines[["left", "bottom"]].set_color("#CCCCCC")

        desde_label = desde_dt.strftime("%d/%m/%Y") if desde_dt else "inicio"
        hasta_label = hasta_dt.strftime("%d/%m/%Y") if hasta_dt else "hoy"
        ax.set_title(f"Consumo de insumos por sucursal \u2014 {desde_label} al {hasta_label}",
                     fontsize=10, color="#333333", pad=8)
        ax.legend(loc="lower right", fontsize=8, framealpha=0.7)
        fig.tight_layout(pad=1.5)
        canvas.draw()

        # Tabla resumen
        gran_toner = gran_unidad = 0
        for suc in sucursales:
            t = suc_data_cache[suc].get("Tóner", 0)
            u = suc_data_cache[suc].get("Unidad Imagen", 0)
            tree_res.insert("", "end", values=(suc, t or "—", u or "—", t + u))
            gran_toner += t; gran_unidad += u
        tree_res.insert("", "end", values=("TOTAL", gran_toner, gran_unidad, gran_toner + gran_unidad),
                        tags=("total",))
        tree_res.tag_configure("total", background="#E3F2FD", font=FONT_BOLD)

    def _ordenar_res(col_idx):
        items = [(tree_res.set(iid, col_idx), iid) for iid in tree_res.get_children("")]
        try:
            items.sort(key=lambda x: (x[0].lstrip("—").strip().isdigit(),
                                      int(x[0].lstrip("—").strip()))
                       if x[0].lstrip("—").strip().isdigit()
                       else (False, x[0].lower()))
        except (ValueError, TypeError):
            items.sort(key=lambda x: str(x[0]).lower())
        # Mantener TOTAL al final
        hijos = tree_res.get_children("")
        total_iid = None
        for iid in hijos:
            if tree_res.item(iid, "values")[0] == "TOTAL":
                total_iid = iid; break
        items = [(v, iid) for v, iid in items if iid != total_iid]
        items.sort(key=lambda x: _sort_key_res(x[0], col_idx))
        for idx, (_, iid) in enumerate(items):
            tree_res.move(iid, "", idx)
        if total_iid:
            tree_res.move(total_iid, "", "end")

    def _sort_key_res(val, col_idx):
        if val == "—":
            return (True, 0)
        try: return (False, int(val))
        except: return (True, str(val).lower())

    def _exportar():
        items = tree_res.get_children()
        if not items:
            messagebox.showwarning("Sin datos", "No hay datos para exportar.", parent=win)
            return
        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"consumo_{datetime.now().strftime('%Y%m%d')}.xlsx", parent=win)
        if not ruta: return
        wb = Workbook(); ws = wb.active; ws.title = "Consumo"
        headers = list(cols_res); widths = [28, 10, 14, 8]
        hf = PatternFill("solid", fgColor="4472C4"); hfont = Font(bold=True, color="FFFFFF")
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=col, value=h); c.fill = hf; c.font = hfont
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[c.column_letter].width = w
        for it in items:
            ws.append(list(tree_res.item(it, "values")))
            if tree_res.item(it, "values")[0] == "TOTAL":
                for col in range(1, len(headers)+1):
                    c = ws.cell(row=ws.max_row, column=col)
                    c.font = Font(bold=True)
                    c.fill = PatternFill("solid", fgColor="E3F2FD")
        wb.save(ruta)
        messagebox.showinfo("Exportado", f"Consumo guardado:\n{ruta}", parent=win)

    def _cargar_detalle(sucursal):
        tree_det.delete(*tree_det.get_children())
        desde_str, hasta_str, _, _ = _parsear_fechas()
        tipo_f = var_tipo.get()
        with db_connect() as conn:
            q = ("SELECT fecha, tipo_insumo, modelo_impresora, cantidad "
                 "FROM envios WHERE sucursal=? AND fecha BETWEEN ? AND ?")
            params = [sucursal, desde_str, hasta_str]
            if tipo_f != "Todos":
                q += " AND tipo_insumo=?"; params.append(tipo_f)
            q += " ORDER BY fecha DESC"
            rows = conn.execute(q, params).fetchall()
        for r in rows:
            try:
                f = datetime.strptime(r[0], "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M")
            except: f = str(r[0] or "")
            tree_det.insert("", "end", values=(f, r[1], r[2], r[3]))
        frame_det.config(text=f"Detalle: {sucursal} ({len(rows)} envíos)")

    def _on_tree_select(event):
        sel = tree_res.selection()
        if not sel: return
        vals = tree_res.item(sel[0], "values")
        if vals and vals[0] not in ("TOTAL", "(sin datos)", "—"):
            _cargar_detalle(vals[0])

    # ── Conexiones ────────────────────────────────────────────────────────────
    btn_aplicar.config(command=_dibujar)
    btn_exportar.config(command=_exportar)
    tree_res.bind("<<TreeviewSelect>>", _on_tree_select)

    var_tipo.trace_add("write", lambda *_: _dibujar())
    var_suc.trace_add("write", lambda *_: _dibujar())

    for i, col in enumerate(cols_res):
        tree_res.heading(col, text=col, command=lambda i=i: _ordenar_res(i))

    win.protocol("WM_DELETE_WINDOW", lambda: (plt.close(fig), win.destroy()))
    _dibujar()
# ---------------------------------------------------------------------------
# Construcción de la interfaz
# ---------------------------------------------------------------------------

def crear_interfaz():
    ventana = tk.Tk()
    ventana.title("Monitoreo de Impresoras")
    ventana.geometry("960x720")
    ventana.config(bg=BG_MAIN)

    # Estilo global ttk
    style = Style()
    style.theme_use("clam")
    style.configure("Treeview",
        background=BG_TREE_PAR, foreground="#333333",
        fieldbackground=BG_TREE_PAR, font=FONT_TABLE, rowheight=26)
    style.configure("Treeview.Heading",
        background=COLOR_ACCENT, foreground="white",
        font=FONT_BOLD, relief="flat")
    style.map("Treeview.Heading",
        background=[("active", COLOR_ACCENT_DARK)])
    style.map("Treeview",
        background=[("selected", "#BFD9F0")],
        foreground=[("selected", "#1A1A1A")])

    evento_cancelar = threading.Event()
    config          = cargar_config()

    # Row 0: Umbrales de alerta + controles auto
    frame_umbrales = tk.Frame(ventana, bg=BG_MAIN)
    frame_umbrales.grid(row=0, column=0, columnspan=4, pady=(10, 5))
    tk.Label(frame_umbrales, text="Nivel bajo:", bg=BG_MAIN, font=FONT_UI).pack(side="left", padx=(10, 2))
    spinbox_bajo = Spinbox(frame_umbrales, from_=1, to=99, width=4)
    spinbox_bajo.delete(0, tk.END)
    spinbox_bajo.insert(0, config.get("umbral_bajo", 10))
    spinbox_bajo.pack(side="left")
    tk.Label(frame_umbrales, text="%", bg=BG_MAIN, font=FONT_UI).pack(side="left", padx=(2, 15))
    tk.Label(frame_umbrales, text="Nivel medio:", bg=BG_MAIN, font=FONT_UI).pack(side="left", padx=(0, 2))
    spinbox_medio = Spinbox(frame_umbrales, from_=1, to=99, width=4)
    spinbox_medio.delete(0, tk.END)
    spinbox_medio.insert(0, config.get("umbral_medio", 25))
    spinbox_medio.pack(side="left")
    tk.Label(frame_umbrales, text="%", bg=BG_MAIN, font=FONT_UI).pack(side="left", padx=(2, 0))
    tk.Label(frame_umbrales, text="  |  ", fg="#aaaaaa", bg=BG_MAIN).pack(side="left")
    var_auto = tk.BooleanVar(value=False)
    tk.Checkbutton(frame_umbrales, text="Auto", variable=var_auto,
                   bg=BG_MAIN, font=FONT_UI, activebackground=BG_MAIN).pack(side="left")
    combo_intervalo = Combobox(frame_umbrales, values=list(INTERVALOS_AUTO.keys()),
                               state="readonly", width=8)
    intervalo_def = config.get("intervalo_auto", "1 hora")
    combo_intervalo.set(intervalo_def)
    combo_intervalo.pack(side="left", padx=(4, 6))
    lbl_proximo = tk.Label(frame_umbrales, text="Próximo: --:--",
                           fg="#555555", bg=BG_MAIN, font=FONT_UI, width=16, anchor="w")
    lbl_proximo.pack(side="left")

    # Row 1: Barra de progreso
    barra_progreso = Progressbar(ventana, orient="horizontal", length=500, mode="determinate")
    barra_progreso.grid(row=1, column=0, columnspan=4, pady=10)

    # Row 2: Barra de búsqueda / filtro
    frame_busqueda = tk.Frame(ventana, bg=BG_MAIN)
    frame_busqueda.grid(row=2, column=0, columnspan=4, padx=10, pady=(0, 4), sticky="ew")
    tk.Label(frame_busqueda, text="Buscar:", bg=BG_MAIN, font=FONT_UI).pack(side="left", padx=(0, 4))
    var_busqueda_sv  = tk.StringVar()
    entrada_busqueda = tk.Entry(frame_busqueda, textvariable=var_busqueda_sv,
                                width=16, font=FONT_UI)
    entrada_busqueda.pack(side="left", padx=(0, 8))
    tk.Label(frame_busqueda, text="Sucursal:", bg=BG_MAIN, font=FONT_UI).pack(side="left", padx=(0, 4))
    var_suc_filtro = tk.StringVar(value="Todas")
    combo_suc_filtro = Combobox(frame_busqueda, textvariable=var_suc_filtro,
                                values=["Todas"] + db_sucursales_activas(),
                                state="readonly", width=14)
    combo_suc_filtro.pack(side="left", padx=(0, 8))
    tk.Label(frame_busqueda, text="Modelo:", bg=BG_MAIN, font=FONT_UI).pack(side="left", padx=(0, 4))
    var_mod_filtro = tk.StringVar(value="Todos")
    combo_mod_filtro = Combobox(frame_busqueda, textvariable=var_mod_filtro,
                                values=["Todos"] + db_modelos_activos(),
                                state="readonly", width=12)
    combo_mod_filtro.pack(side="left", padx=(0, 8))
    var_solo_alertas = tk.BooleanVar(value=False)
    tk.Checkbutton(frame_busqueda, text="Solo alertas", variable=var_solo_alertas,
                   bg=BG_MAIN, font=FONT_UI, activebackground=BG_MAIN).pack(side="left")

    # Row 3: Tabla de resultados (6 columnas)
    ventana.columnconfigure(0, weight=1)
    ventana.rowconfigure(3, weight=1)

    frame_tree = tk.Frame(ventana, bg=BG_MAIN)
    frame_tree.grid(row=3, column=0, columnspan=4, padx=10, pady=0, sticky="nsew")
    frame_tree.columnconfigure(0, weight=1)
    frame_tree.rowconfigure(0, weight=1)

    col_widths_tree = {
        "Sucursal": 110, "IP": 120, "Modelo": 140, "Último Monitoreo": 130,
        "Tóner (%)": 90, "Unidad Imagen (%)": 110, "Kit Mantenimiento (%)": 125,
    }
    tree = Treeview(frame_tree, columns=COLUMNAS_TREE, show="headings", height=10)
    for col in COLUMNAS_TREE:
        tree.heading(col, text=col)
        tree.column(col, anchor="center", width=col_widths_tree.get(col, 100))
    # Tags: par/impar primero (menor prioridad), luego alertas (mayor prioridad)
    tree.tag_configure("par",      background=BG_TREE_PAR)
    tree.tag_configure("impar",    background=BG_TREE_IMPAR)
    tree.tag_configure("sin_datos", background=COLOR_SIN_DATOS, foreground="#555555")
    tree.tag_configure("medio",    background=COLOR_MEDIO,     foreground="#5A4500")
    tree.tag_configure("bajo",     background=COLOR_BAJO,      foreground="#7A0000")
    scrollbar = Scrollbar(frame_tree, orient="vertical", command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    tree.grid(row=0, column=0, sticky="nsew")
    scrollbar.grid(row=0, column=1, sticky="ns")

    # Row 4: Panel de resumen
    frame_resumen = tk.LabelFrame(ventana, text="Resumen",
                                  bg=BG_MAIN, font=FONT_BOLD, padx=10, pady=5)
    frame_resumen.grid(row=4, column=0, columnspan=4, padx=10, pady=(5, 0), sticky="ew")
    labels_resumen = {}
    for texto, clave, fila, col in [
        ("Total:",        "total",        0, 0),
        ("Respondieron:", "respondieron", 0, 2),
        ("Sin datos:",    "sin_datos",    0, 4),
        ("Nivel bajo:",   "bajo",         1, 0),
        ("Nivel medio:",  "medio",        1, 2),
    ]:
        tk.Label(frame_resumen, text=texto, bg=BG_MAIN, font=FONT_BOLD).grid(
            row=fila, column=col, padx=(10, 2), pady=2, sticky="e")
        labels_resumen[clave] = tk.Label(frame_resumen, text="\u2014",
                                         width=4, bg=BG_MAIN, font=FONT_NUM)
        labels_resumen[clave].grid(row=fila, column=col + 1, pady=2, sticky="w")

    # Fila 2: Stock depósito
    tk.Label(frame_resumen, text="Stock crítico:", bg=BG_MAIN, font=FONT_BOLD).grid(
        row=2, column=0, padx=(10, 2), pady=(4, 6), sticky="e")
    lbl_stock_crit = tk.Label(frame_resumen, text="\u2014", width=4,
                               bg=BG_MAIN, font=FONT_NUM, fg=COLOR_BAJO)
    lbl_stock_crit.grid(row=2, column=1, pady=(4, 6), sticky="w")
    labels_resumen["stock_critico"] = lbl_stock_crit

    tk.Label(frame_resumen, text="Stock bajo:", bg=BG_MAIN, font=FONT_BOLD).grid(
        row=2, column=2, padx=(10, 2), pady=(4, 6), sticky="e")
    lbl_stock_bajo = tk.Label(frame_resumen, text="\u2014", width=4,
                               bg=BG_MAIN, font=FONT_NUM, fg=COLOR_MEDIO)
    lbl_stock_bajo.grid(row=2, column=3, pady=(4, 6), sticky="w")
    labels_resumen["stock_bajo"] = lbl_stock_bajo

    # Fila 3: Último monitoreo
    tk.Label(frame_resumen, text="Último monitoreo:", bg=BG_MAIN, font=FONT_BOLD).grid(
        row=3, column=0, padx=(10, 2), pady=(2, 2), sticky="e")
    labels_resumen["fecha_ult"] = tk.Label(frame_resumen, text="\u2014",
                                           bg=BG_MAIN, font=FONT_UI, fg="gray")
    labels_resumen["fecha_ult"].grid(row=3, column=1, columnspan=3, pady=(2, 2), sticky="w")

    # Fila 4: Mini dashboard stock
    labels_resumen["stock_mini"] = tk.Label(frame_resumen, text="",
                                            bg=BG_MAIN, font=("Segoe UI", 7), fg="#666666",
                                            anchor="w", wraplength=500, justify="left")
    labels_resumen["stock_mini"].grid(row=4, column=0, columnspan=6, pady=(0, 2), sticky="ew")

    # Row 5: Botones de monitoreo
    frame_btns1 = tk.Frame(ventana, bg=BG_MAIN)
    frame_btns1.grid(row=5, column=0, columnspan=4, pady=(8, 2))
    btn_iniciar  = tk.Button(frame_btns1, text="Iniciar Monitoreo")
    btn_iniciar.pack(side="left", padx=6)
    btn_cancelar = tk.Button(frame_btns1, text="Cancelar", state="disabled")
    btn_cancelar.pack(side="left", padx=6)
    btn_refrescar = tk.Button(frame_btns1, text="Refrescar")
    btn_refrescar.pack(side="left", padx=6)
    btn_grafico  = tk.Button(frame_btns1, text="Ver Gráfico", state="disabled")
    btn_grafico.pack(side="left", padx=6)
    btn_exportar = tk.Button(frame_btns1, text="Exportar Excel", state="disabled")
    btn_exportar.pack(side="left", padx=6)

    # Row 6: Botones de gestión
    frame_btns2 = tk.Frame(ventana, bg=BG_MAIN)
    frame_btns2.grid(row=6, column=0, columnspan=4, pady=(2, 6))
    btn_catalogo  = tk.Button(frame_btns2, text="Impresoras")
    btn_catalogo.pack(side="left", padx=6)
    btn_historial = tk.Button(frame_btns2, text="Historial")
    btn_historial.pack(side="left", padx=6)
    btn_insumos   = tk.Button(frame_btns2, text="Envío de Insumos")
    btn_insumos.pack(side="left", padx=6)
    btn_stock     = tk.Button(frame_btns2, text="Stock Depósito")
    btn_stock.pack(side="left", padx=6)
    btn_estadist  = tk.Button(frame_btns2, text="Estadísticas")
    btn_estadist.pack(side="left", padx=6)
    btn_email     = tk.Button(frame_btns2, text="Configuración")
    btn_email.pack(side="left", padx=6)

    # Row 7: Etiqueta de estado
    resultado_label = tk.Label(ventana, text="", fg="green", bg=BG_MAIN, font=FONT_UI)
    resultado_label.grid(row=7, column=0, columnspan=4, pady=(0, 10))

    # Construir contexto con todas las referencias a widgets
    ctx = SimpleNamespace(
        ventana=ventana,
        barra_progreso=barra_progreso,
        tree=tree,
        btn_iniciar=btn_iniciar,
        btn_cancelar=btn_cancelar,
        btn_grafico=btn_grafico,
        btn_exportar=btn_exportar,
        spinbox_bajo=spinbox_bajo,
        spinbox_medio=spinbox_medio,
        labels_resumen=labels_resumen,
        resultado_label=resultado_label,
        evento_cancelar=evento_cancelar,
        entrada_busqueda=entrada_busqueda,
        var_solo_alertas=var_solo_alertas,
        var_auto=var_auto,
        combo_intervalo=combo_intervalo,
        lbl_proximo=lbl_proximo,
        after_id=None,
        filas_tabla=[],
        sort_col=None,
        sort_asc=True,
        var_suc_filtro=var_suc_filtro,
        var_mod_filtro=var_mod_filtro,
    )

    # Asignar comandos ahora que ctx está construido
    btn_iniciar.config(command=lambda:  iniciar_monitoreo(ctx))
    btn_cancelar.config(command=ctx.evento_cancelar.set)
    btn_grafico.config(command=lambda:  ver_grafico(ctx))
    btn_exportar.config(command=lambda: exportar_excel(ctx))
    btn_catalogo.config(command=abrir_catalogo_impresoras)
    btn_historial.config(command=abrir_historial)
    btn_insumos.config(command=abrir_envio_insumos)
    btn_stock.config(command=abrir_stock_deposito)
    btn_estadist.config(command=abrir_estadisticas_consumo)
    btn_email.config(command=abrir_configuracion)

    # Aplicar estilo a botones
    _estilo_btn(btn_iniciar,  primario=True)
    _estilo_btn(btn_exportar, primario=True)
    _estilo_btn(btn_cancelar, primario=False)
    _estilo_btn(btn_refrescar, primario=False)
    _estilo_btn(btn_grafico,  primario=False)
    _estilo_btn(btn_catalogo, primario=False)
    _estilo_btn(btn_historial,primario=False)
    _estilo_btn(btn_insumos,  primario=False)
    _estilo_btn(btn_stock,    primario=False)
    _estilo_btn(btn_estadist, primario=False)
    _estilo_btn(btn_email,    primario=False)

    # Tooltips
    Tooltip(btn_iniciar,  "Consultar todas las impresoras activas")
    Tooltip(btn_cancelar, "Detener el monitoreo en curso")
    Tooltip(btn_refrescar, "Recargar el último monitoreo desde la base de datos")
    Tooltip(btn_grafico,  "Ver historial gráfico de la impresora seleccionada")
    Tooltip(btn_exportar, "Exportar resultados a Excel (.xlsx)")
    Tooltip(btn_catalogo, "Gestionar el catálogo de impresoras")
    Tooltip(btn_historial,"Ver historial completo de monitoreos")
    Tooltip(btn_insumos,  "Registrar y consultar envíos de insumos")
    Tooltip(btn_stock,    "Gestionar stock de insumos en depósito")
    Tooltip(btn_estadist, "Ver estadísticas de consumo de insumos por sucursal")
    Tooltip(btn_email,    "Configurar ruta de base de datos y notificaciones por email")

    # Toggle monitoreo automático
    def _toggle_auto():
        if ctx.var_auto.get():
            _iniciar_auto(ctx)
        else:
            _cancelar_auto(ctx)
    var_auto.trace_add("write", lambda *_: _toggle_auto())

    # Encabezados de columna con ordenamiento
    for i, col in enumerate(COLUMNAS_TREE):
        tree.heading(col, text=col, command=lambda i=i: ordenar_por_columna(ctx, i))

    # Filtros en tiempo real
    var_busqueda_sv.trace_add("write",  lambda *_: aplicar_filtro(ctx))
    var_solo_alertas.trace_add("write", lambda *_: aplicar_filtro(ctx))
    var_suc_filtro.trace_add("write",  lambda *_: aplicar_filtro(ctx))
    var_mod_filtro.trace_add("write",  lambda *_: aplicar_filtro(ctx))

    # Habilitar/deshabilitar "Ver Gráfico" según la selección en la tabla
    tree.bind("<<TreeviewSelect>>",
              lambda _: ctx.btn_grafico.config(
                  state="normal" if ctx.tree.selection() else "disabled"))

    # Menú contextual (clic derecho)
    menu_context = tk.Menu(ventana, tearoff=0)
    menu_context.add_command(label="Ver Gráfico",
                             command=lambda: ver_grafico(ctx) if tree.selection() else None)
    menu_context.add_command(label="Copiar IP",
                             command=lambda: _copiar_ip(ctx) if tree.selection() else None)
    menu_context.add_command(label="Abrir en Catálogo",
                             command=lambda: _abrir_en_catalogo(ctx) if tree.selection() else None)

    def _mostrar_contexto(event):
        iid = tree.identify_row(event.y)
        if iid:
            tree.selection_set(iid)
            menu_context.tk_popup(event.x_root, event.y_root)

    tree.bind("<Button-3>", _mostrar_contexto)

    def _copiar_ip(ctx_):
        sel = ctx_.tree.selection()
        if sel:
            ip = ctx_.tree.item(sel[0], "values")[1]
            ventana.clipboard_clear()
            ventana.clipboard_append(ip)
            ctx_.resultado_label.config(text=f"IP copiada: {ip}", fg="gray")

    def _abrir_en_catalogo(ctx_):
        sel = ctx_.tree.selection()
        if sel:
            ip = ctx_.tree.item(sel[0], "values")[1]
            abrir_catalogo_impresoras(seleccionar_ip=ip)

    def _refrescar():
        u_bajo  = int(ctx.spinbox_bajo.get())
        u_medio = int(ctx.spinbox_medio.get())
        filas, fecha = db_cargar_ultimo_monitoreo(u_bajo, u_medio)
        ctx.tree.delete(*ctx.tree.get_children())
        ctx.filas_tabla.clear()
        ctx.sort_col = None
        ctx.sort_asc = True
        for col in COLUMNAS_TREE:
            ctx.tree.heading(col, text=col)
        if filas:
            ctx.filas_tabla.extend(filas)
            for idx, fila in enumerate(filas):
                parity = "par" if idx % 2 == 0 else "impar"
                tags   = (fila[7],) if fila[7] else (parity,)
                ctx.tree.insert("", "end", values=fila[:7], tags=tags)
            n_bajo = n_medio = n_sin_datos = 0
            for f in filas:
                if   f[7] == "sin_datos": n_sin_datos += 1
                elif f[7] == "bajo":      n_bajo      += 1
                elif f[7] == "medio":     n_medio     += 1
            total = len(filas)
            respondieron = total - n_sin_datos
            _actualizar_resumen(ctx, total, respondieron, n_sin_datos, n_bajo, n_medio, fecha)
            try:
                ff = datetime.strptime(fecha, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M")
            except: ff = str(fecha)
            ctx.resultado_label.config(text=f"Último monitoreo: {ff}", fg="gray")
            ctx.btn_exportar.config(state="normal")
            _estilo_btn(ctx.btn_exportar, primario=True)
        else:
            ctx.resultado_label.config(text="No hay datos de monitoreo", fg="gray")
        _actualizar_stock_resumen(ctx)

    btn_refrescar.config(command=_refrescar)

    # Atajos de teclado
    def _key_shortcuts(event):
        if event.keysym == "F5" or (event.state & 0x4 and event.keysym.lower() == "r"):
            if ctx.btn_iniciar["state"] == "normal":
                iniciar_monitoreo(ctx)
            return "break"
        if event.state & 0x4 and event.keysym.lower() == "f":
            entrada_busqueda.focus_set()
            return "break"
        if event.state & 0x4 and event.keysym.lower() == "e":
            if ctx.btn_exportar["state"] == "normal":
                exportar_excel(ctx)
            return "break"
        if event.state & 0x4 and event.keysym.lower() == "g":
            if ctx.btn_grafico["state"] == "normal":
                ver_grafico(ctx)
            return "break"
        if event.keysym == "Escape":
            if ctx.btn_cancelar["state"] == "normal":
                ctx.evento_cancelar.set()
            return "break"
        return None

    ventana.bind("<Key>", _key_shortcuts)

    # Cargar el último monitoreo en la tabla al abrir la aplicación
    filas_inicio, fecha_inicio = db_cargar_ultimo_monitoreo(
        config.get("umbral_bajo", 10), config.get("umbral_medio", 25))
    if filas_inicio:
        ctx.filas_tabla.extend(filas_inicio)
        for idx, fila in enumerate(filas_inicio):
            parity = "par" if idx % 2 == 0 else "impar"
            tags   = (fila[7],) if fila[7] else (parity,)
            tree.insert("", "end", values=fila[:7], tags=tags)
        n_bajo = n_medio = n_sin_datos = 0
        for fila in filas_inicio:
            if   fila[7] == "sin_datos": n_sin_datos += 1
            elif fila[7] == "bajo":      n_bajo      += 1
            elif fila[7] == "medio":     n_medio     += 1
        total        = len(filas_inicio)
        respondieron = total - n_sin_datos
        _actualizar_resumen(ctx, total, respondieron, n_sin_datos, n_bajo, n_medio, fecha_inicio)
        try:
            fecha_fmt = datetime.strptime(fecha_inicio, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M")
        except (ValueError, TypeError):
            fecha_fmt = str(fecha_inicio)
        resultado_label.config(text=f"Último monitoreo: {fecha_fmt}", fg="gray")
        btn_exportar.config(state="normal")
        _estilo_btn(btn_exportar, primario=True)

    # Actualizar dashboard de stock al iniciar
    _actualizar_stock_resumen(ctx)

    ventana.mainloop()


_inicializar_db_path()
init_db()
crear_interfaz()
